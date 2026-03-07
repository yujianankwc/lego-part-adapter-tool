import csv
import io
import json
import os
import re
import threading
import time
import uuid
from copy import deepcopy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.error import HTTPError
from urllib.parse import quote
from urllib.request import Request, urlopen

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
PART_ADAPTER_DIR = DATA_DIR / 'part_adapter'
RULES_PATH = PART_ADAPTER_DIR / 'rules.json'
JOBS_PATH = PART_ADAPTER_DIR / 'jobs.json'
ANALYTICS_PATH = PART_ADAPTER_DIR / 'analytics.json'

REBRICKABLE_COLOR_NAME_ZH: Dict[str, str] = {
    'White': '白色',
    'Black': '黑色',
    'Red': '红色',
    'Blue': '蓝色',
    'Yellow': '黄色',
    'Green': '绿色',
    'Tan': '棕褐',
    'Dark Tan': '深棕褐',
    'Light Gray': '浅灰',
    'Dark Gray': '深灰',
    'Light Bluish Gray': '浅蓝灰',
    'Dark Bluish Gray': '深蓝灰',
    'Reddish Brown': '红棕',
    'Dark Brown': '深棕',
    'Dark Orange': '深橙',
    'Orange': '橙色',
    'Bright Light Orange': '浅亮橙',
    'Medium Nougat': '中肉色',
    'Light Nougat': '浅肉色',
    'Nougat': '肉色',
    'Brick Yellow': '砖黄',
    'Bright Light Yellow': '亮浅黄',
    'Lime': '青柠绿',
    'Bright Green': '亮绿色',
    'Yellowish Green': '黄绿色',
    'Olive Green': '橄榄绿',
    'Dark Green': '深绿',
    'Sand Green': '沙绿',
    'Dark Turquoise': '深青绿',
    'Medium Azure': '中碧蓝',
    'Dark Azure': '深天蓝',
    'Medium Blue': '中蓝',
    'Bright Light Blue': '亮浅蓝',
    'Sand Blue': '沙蓝',
    'Dark Blue': '深蓝',
    'Dark Purple': '深紫',
    'Medium Lavender': '中薰衣草紫',
    'Lavender': '薰衣草紫',
    'Light Purple': '浅紫',
    'Bright Pink': '亮粉',
    'Dark Pink': '深粉',
    'Magenta': '洋红',
    'Dark Red': '深红',
    'Light Pink': '浅粉',
    'Coral': '珊瑚色',
    'Flat Silver': '浅银灰',
    'Pearl Gold': '珍珠金',
    'Warm Gold': '暖金',
    'Cool Silver': '冷银',
    'Metallic Silver': '金属银',
    'Trans-Clear': '透明',
    'Trans-Red': '透明红',
    'Trans-Dark Pink': '透明深粉',
    'Trans-Orange': '透明橙',
    'Trans-Yellow': '透明黄',
    'Trans-Green': '透明绿',
    'Trans-Neon Green': '透明荧光绿',
    'Trans-Dark Blue': '透明深蓝',
    'Trans-Light Blue': '透明浅蓝',
    'Trans-Purple': '透明紫',
    'Trans-Brown': '透明棕',
    'Glow In Dark Opaque': '夜光不透明',
    'Glow In Dark Trans': '夜光透明',
}


def _now_iso() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


DEFAULT_RULES: Dict[str, Any] = {
    'updated_at': '',
    'gobricks_sync_meta': {
        'last_sync_at': '',
        'last_sync_count': 0,
        'last_sync_start_time': '',
        'last_sync_end_time': '',
        'base_url': 'https://api.gobricks.cn',
        'need_detail_info': True,
    },
    'gobricks_item_index': {},
    'gobricks_category_index': {},
    'rebrickable_part_catalog': {},
    'rebrickable_part_category_catalog': {},
    'element_part_map': {},
    'print_part_map': {},
    'external_part_map': {},
    'part_detail_cache': {},
    'part_image_cache': {},
    'gobricks_color_catalog': {
        '010': {'name': '大红', 'name_en': 'Red', 'finish': 'solid', 'swatch_hex': '#ef5349'},
        '011': {'name': '粉红', 'name_en': 'Bright Pink', 'finish': 'solid', 'swatch_hex': '#e85c9a'},
        '012': {'name': '桃红', 'name_en': 'Dark Pink', 'finish': 'solid', 'swatch_hex': '#d44686'},
        '013': {'name': '紫红', 'name_en': 'Magenta', 'finish': 'solid', 'swatch_hex': '#b1467d'},
        '014': {'name': '梅红', 'name_en': 'Dark Red', 'finish': 'solid', 'swatch_hex': '#7f1f2d'},
        '017': {'name': '浅粉', 'name_en': 'Light Pink', 'finish': 'solid', 'swatch_hex': '#f4b6cf'},
        '018': {'name': '珊瑚橘', 'name_en': 'Coral', 'finish': 'solid', 'swatch_hex': '#f17f6b'},
        '021': {'name': '橙黄', 'name_en': 'Orange', 'finish': 'solid', 'swatch_hex': '#f28a2e'},
        '027': {'name': '白黄', 'name_en': 'Light Yellow', 'finish': 'solid', 'swatch_hex': '#f1dd9a'},
        '029': {'name': '浅黄', 'name_en': 'Light Yellow', 'finish': 'solid', 'swatch_hex': '#efe1a7'},
        '030': {'name': '黄色', 'name_en': 'Yellow', 'finish': 'solid', 'swatch_hex': '#f0c21b'},
        '031': {'name': '杏黄', 'name_en': 'Tan', 'finish': 'solid', 'swatch_hex': '#cfa867'},
        '032': {'name': '浅肉', 'name_en': 'Light Nougat', 'finish': 'solid', 'swatch_hex': '#f1c198'},
        '033': {'name': '牙黄', 'name_en': 'Bright Light Yellow', 'finish': 'solid', 'swatch_hex': '#f1dea1'},
        '034': {'name': '土黄', 'name_en': 'Dark Tan', 'finish': 'solid', 'swatch_hex': '#b68d5b'},
        '035': {'name': '金黄', 'name_en': 'Pearl Gold', 'finish': 'metallic', 'swatch_hex': '#b48f3f'},
        '036': {'name': '深黄', 'name_en': 'Bright Light Orange', 'finish': 'solid', 'swatch_hex': '#f3b54b'},
        '037': {'name': '亮金', 'name_en': 'Bright Gold', 'finish': 'metallic', 'swatch_hex': '#c99843'},
        '038': {'name': '中肉', 'name_en': 'Nougat', 'finish': 'solid', 'swatch_hex': '#d59a63'},
        '040': {'name': '绿色', 'name_en': 'Green', 'finish': 'solid', 'swatch_hex': '#357f3c'},
        '041': {'name': '军绿', 'name_en': 'Dark Green', 'finish': 'solid', 'swatch_hex': '#35502f'},
        '042': {'name': '青绿', 'name_en': 'Lime', 'finish': 'solid', 'swatch_hex': '#95c93d'},
        '043': {'name': '果绿', 'name_en': 'Bright Green', 'finish': 'solid', 'swatch_hex': '#49b244'},
        '044': {'name': '浅绿', 'name_en': 'Yellowish Green', 'finish': 'solid', 'swatch_hex': '#b8d48c'},
        '045': {'name': '粉绿', 'name_en': 'Light Aqua', 'finish': 'solid', 'swatch_hex': '#b8d4d2'},
        '046': {'name': '翠绿', 'name_en': 'Medium Azure', 'finish': 'solid', 'swatch_hex': '#3ea3b9'},
        '047': {'name': '深绿', 'name_en': 'Dark Green', 'finish': 'solid', 'swatch_hex': '#0c6b47'},
        '048': {'name': '清绿', 'name_en': 'Sand Green', 'finish': 'solid', 'swatch_hex': '#84a58e'},
        '049': {'name': '马丁绿', 'name_en': 'Olive Green', 'finish': 'solid', 'swatch_hex': '#8a8f4f'},
        '050': {'name': '蓝色', 'name_en': 'Blue', 'finish': 'solid', 'swatch_hex': '#1e66b2'},
        '051': {'name': '天蓝', 'name_en': 'Dark Azure', 'finish': 'solid', 'swatch_hex': '#2a86be'},
        '052': {'name': '浅蓝', 'name_en': 'Medium Blue', 'finish': 'solid', 'swatch_hex': '#66a7d9'},
        '053': {'name': '粉蓝', 'name_en': 'Bright Light Blue', 'finish': 'solid', 'swatch_hex': '#96c9ea'},
        '054': {'name': '蓝灰', 'name_en': 'Sand Blue', 'finish': 'solid', 'swatch_hex': '#7693b1'},
        '055': {'name': '宝蓝', 'name_en': 'Dark Blue', 'finish': 'solid', 'swatch_hex': '#274f91'},
        '056': {'name': '浅灰蓝', 'name_en': 'Blue Gray', 'finish': 'solid', 'swatch_hex': '#8d97b6'},
        '060': {'name': '紫色', 'name_en': 'Dark Purple', 'finish': 'solid', 'swatch_hex': '#6a478f'},
        '062': {'name': '蓝光紫', 'name_en': 'Medium Lavender', 'finish': 'solid', 'swatch_hex': '#9e7dc1'},
        '063': {'name': '浅紫', 'name_en': 'Lavender', 'finish': 'solid', 'swatch_hex': '#b49acb'},
        '064': {'name': '粉紫', 'name_en': 'Light Purple', 'finish': 'solid', 'swatch_hex': '#ceb8d6'},
        '071': {'name': '浅灰', 'name_en': 'Light Bluish Gray', 'finish': 'solid', 'swatch_hex': '#b9b9b6'},
        '072': {'name': '深灰', 'name_en': 'Dark Bluish Gray', 'finish': 'solid', 'swatch_hex': '#6f7274'},
        '073': {'name': '浅银灰', 'name_en': 'Flat Silver', 'finish': 'metallic', 'swatch_hex': '#a5a8aa'},
        '080': {'name': '黑色', 'name_en': 'Black', 'finish': 'solid', 'swatch_hex': '#1f2328'},
        '081': {'name': '棕色', 'name_en': 'Reddish Brown', 'finish': 'solid', 'swatch_hex': '#6f3718'},
        '082': {'name': '棕玉', 'name_en': 'Dark Brown', 'finish': 'solid', 'swatch_hex': '#4d2a14'},
        '083': {'name': '咖啡', 'name_en': 'Dark Orange', 'finish': 'solid', 'swatch_hex': '#a85f22'},
        '084': {'name': '浅咖啡', 'name_en': 'Medium Nougat', 'finish': 'solid', 'swatch_hex': '#c48a4a'},
        '085': {'name': '橄榄', 'name_en': 'Olive', 'finish': 'solid', 'swatch_hex': '#7c7442'},
        '090': {'name': '白色', 'name_en': 'White', 'finish': 'solid', 'swatch_hex': '#f2f3f2'},
        '091': {'name': '乳白', 'name_en': 'Ivory White', 'finish': 'solid', 'swatch_hex': '#f2efe4'},
        '092': {'name': '灰白', 'name_en': 'Warm White', 'finish': 'solid', 'swatch_hex': '#e5e1d8'},
        '110': {'name': '明红', 'name_en': 'Trans-Red', 'finish': 'transparent', 'swatch_hex': '#d96a62'},
        '111': {'name': '暗紫红', 'name_en': 'Trans-Dark Pink', 'finish': 'transparent', 'swatch_hex': '#c98a9f'},
        '120': {'name': '明橙', 'name_en': 'Trans-Orange', 'finish': 'transparent', 'swatch_hex': '#efb04f'},
        '130': {'name': '明黄', 'name_en': 'Trans-Yellow', 'finish': 'transparent', 'swatch_hex': '#e5dc6d'},
        '140': {'name': '明绿', 'name_en': 'Trans-Green', 'finish': 'transparent', 'swatch_hex': '#8cc47c'},
        '141': {'name': '明草绿', 'name_en': 'Trans-Neon Green', 'finish': 'transparent', 'swatch_hex': '#b8e56d'},
        '150': {'name': '明蓝', 'name_en': 'Trans-Dark Blue', 'finish': 'transparent', 'swatch_hex': '#7da9cc'},
        '152': {'name': '明浅蓝', 'name_en': 'Trans-Light Blue', 'finish': 'transparent', 'swatch_hex': '#b7d3e6'},
        '160': {'name': '明紫', 'name_en': 'Trans-Purple', 'finish': 'transparent', 'swatch_hex': '#b39ac9'},
        '170': {'name': '明棕', 'name_en': 'Trans-Brown', 'finish': 'transparent', 'swatch_hex': '#aea39a'},
        '180': {'name': '全透明', 'name_en': 'Trans-Clear', 'finish': 'transparent', 'swatch_hex': '#e9eef4'},
        '243': {'name': '玉绿', 'name_en': 'Dark Turquoise', 'finish': 'solid', 'swatch_hex': '#2f8a90'},
    },
    'exact_combo_map': {},
    'shortage_combo_map': {},
    'lego_color_catalog': {
        '1': {'name': '白色', 'name_en': 'White', 'family': 'neutral', 'brightness': 96, 'warmth': 50, 'finish': 'solid'},
        '2': {'name': '', 'name_en': 'Light Gray', 'family': 'gray', 'brightness': 84, 'warmth': 48, 'finish': 'solid'},
        '3': {'name': '', 'name_en': 'Dark Gray', 'family': 'gray', 'brightness': 38, 'warmth': 48, 'finish': 'solid'},
        '5': {'name': '', 'name_en': 'Brick Yellow', 'family': 'tan', 'brightness': 72, 'warmth': 70, 'finish': 'solid'},
        '9': {'name': '', 'name_en': '', 'family': 'purple', 'brightness': 74, 'warmth': 46, 'finish': 'solid'},
        '10': {'name': '', 'name_en': '', 'family': 'green', 'brightness': 62, 'warmth': 48, 'finish': 'solid'},
        '11': {'name': '黑色', 'name_en': 'Black', 'family': 'neutral', 'brightness': 8, 'warmth': 50, 'finish': 'solid'},
        '21': {'name': '红色', 'name_en': 'Red', 'family': 'red', 'brightness': 48, 'warmth': 92, 'finish': 'solid'},
        '24': {'name': '黄色', 'name_en': 'Yellow', 'family': 'yellow', 'brightness': 85, 'warmth': 90, 'finish': 'solid'},
        '55': {'name': '', 'name_en': '', 'family': 'blue', 'brightness': 56, 'warmth': 40, 'finish': 'solid'},
        '71': {'name': '浅灰', 'name_en': 'Light Gray', 'family': 'gray', 'brightness': 76, 'warmth': 46, 'finish': 'solid'},
        '84': {'name': '浅金', 'name_en': 'Light Tan', 'family': 'tan', 'brightness': 72, 'warmth': 72, 'finish': 'solid'},
        '86': {'name': '浅蓝灰', 'name_en': 'Light Bluish Gray', 'family': 'gray', 'brightness': 72, 'warmth': 42, 'finish': 'solid'},
        '85': {'name': '', 'name_en': '', 'family': 'gray', 'brightness': 42, 'warmth': 44, 'finish': 'solid'},
        '89': {'name': '', 'name_en': '', 'family': 'purple', 'brightness': 30, 'warmth': 44, 'finish': 'solid'},
        '88': {'name': '红棕', 'name_en': 'Reddish Brown', 'family': 'brown', 'brightness': 30, 'warmth': 84, 'finish': 'solid'},
        '90': {'name': '中肉色', 'name_en': 'Medium Nougat', 'family': 'tan', 'brightness': 68, 'warmth': 75, 'finish': 'solid'},
        '102': {'name': '透明蓝', 'name_en': 'Trans Blue', 'family': 'blue', 'brightness': 68, 'warmth': 35, 'finish': 'transparent'},
        '105': {'name': '', 'name_en': '', 'family': 'blue', 'brightness': 82, 'warmth': 36, 'finish': 'solid'},
        '150': {'name': '浅蓝灰', 'name_en': 'Light Bluish Gray', 'family': 'gray', 'brightness': 74, 'warmth': 42, 'finish': 'solid'},
        '167': {'name': '', 'name_en': '', 'family': 'gray', 'brightness': 68, 'warmth': 48, 'finish': 'solid'},
        '194': {'name': '暖灰', 'name_en': 'Medium Stone Gray', 'family': 'gray', 'brightness': 58, 'warmth': 56, 'finish': 'solid'},
        '199': {'name': '深暖灰', 'name_en': 'Dark Stone Gray', 'family': 'gray', 'brightness': 44, 'warmth': 58, 'finish': 'solid'},
        '241': {'name': '棕褐', 'name_en': 'Dark Tan', 'family': 'brown', 'brightness': 34, 'warmth': 82, 'finish': 'solid'},
        '297': {'name': '透明浅蓝', 'name_en': 'Trans Light Blue', 'family': 'blue', 'brightness': 78, 'warmth': 36, 'finish': 'transparent'},
        '312': {'name': '深蓝灰', 'name_en': 'Dark Bluish Gray', 'family': 'gray', 'brightness': 48, 'warmth': 40, 'finish': 'solid'},
    },
    'exact_part_map': {
        '3001': 'G3001',
        '3020': 'G3020',
        '3068b': 'G3068B',
        '50950': 'G50950',
    },
    'part_alias_map': {
        '4589': '59900',
    },
    'color_rules': {
        '1': {
            'safe': {'to': '1', 'note': '基础白色可直接对应', 'risk': 'A'},
            'balanced': {'to': '1', 'note': '基础白色可直接对应', 'risk': 'A'},
            'aggressive': {'to': '1', 'note': '基础白色可直接对应', 'risk': 'A'},
        },
        '21': {
            'safe': {'to': '21', 'note': '基础红色可直接对应', 'risk': 'A'},
            'balanced': {'to': '21', 'note': '基础红色可直接对应', 'risk': 'A'},
            'aggressive': {'to': '21', 'note': '基础红色可直接对应', 'risk': 'A'},
        },
        '24': {
            'safe': {'to': '24', 'note': '基础黄色可直接对应', 'risk': 'A'},
            'balanced': {'to': '24', 'note': '基础黄色可直接对应', 'risk': 'A'},
            'aggressive': {'to': '24', 'note': '基础黄色可直接对应', 'risk': 'A'},
        },
        '194': {
            'safe': None,
            'balanced': {'to': '199', 'note': '高砖无原色，替换为相近暖灰', 'risk': 'B'},
            'aggressive': {'to': '199', 'note': '高砖无原色，替换为相近暖灰', 'risk': 'B'},
        },
        '297': {
            'safe': None,
            'balanced': None,
            'aggressive': {'to': '102', 'note': '激进模式下使用相近透明蓝', 'risk': 'C'},
        },
    },
    'substitutions': {
        '3039': {
            'display': {'to': 'G50950', 'risk': 'B', 'note': '用近似斜坡件替代，展示面差异可接受'},
        },
        '2431': {
            'display': {'to': 'G61252', 'risk': 'C', 'note': '连接件替换可能影响咬合，需人工确认'},
            'structural': {'to': 'G61252', 'risk': 'C', 'note': '结构件连接替代风险较高，默认关闭'},
        },
    },
    'part_meta': {
        '3001': {'family': 'brick', 'structural': True},
        '3020': {'family': 'plate', 'structural': True},
        '3068b': {'family': 'tile', 'structural': False},
        '50950': {'family': 'slope', 'structural': False},
        '3039': {'family': 'slope', 'structural': False},
        '2431': {'family': 'clip', 'structural': True},
        '973pb001': {'family': 'printed', 'structural': False, 'printed': True},
    },
    'source_refs': [
        {
            'id': 'bricklink-color-guide',
            'name': 'BrickLink Color Guide',
            'domain': 'bricklink.com',
            'url': 'https://v2.bricklink.com/en-us/catalog/color-guide',
            'source_type': 'color-reference',
            'notes': '用于人工核对 BrickLink 颜色 ID、LEGO 对应色名与时间线。',
        },
        {
            'id': 'bricklink-catalog',
            'name': 'BrickLink Catalog',
            'domain': 'bricklink.com',
            'url': 'https://www.bricklink.com/catalog.asp',
            'source_type': 'part-reference',
            'notes': '用于人工查询零件编号、颜色可用性与目录分类。',
        },
        {
            'id': 'brickset-colours',
            'name': 'Brickset Colours',
            'domain': 'brickset.com',
            'url': 'https://brickset.com/colours',
            'source_type': 'color-reference',
            'notes': '用于查看颜色家族、在产/停产状态，并参考 BrickLink 命名映射。',
        },
        {
            'id': 'rebrickable-api-docs',
            'name': 'Rebrickable API Docs',
            'domain': 'rebrickable.com',
            'url': 'https://rebrickable.com/api/v3/docs/',
            'source_type': 'api-docs',
            'notes': '官方 API 文档；适合后续接入颜色、零件和 CSV 下载能力，需 API key。',
        },
        {
            'id': 'rebrickable-colors',
            'name': 'Rebrickable Colors',
            'domain': 'rebrickable.com',
            'url': 'https://rebrickable.com/colors/',
            'source_type': 'color-reference',
            'notes': '可核对 RGB、LEGO / BrickLink / BrickOwl 颜色映射字段。',
        },
        {
            'id': 'mygobricks-bulk-tool',
            'name': 'MyGoBricks Bulk Tool',
            'domain': 'mygobricks.com',
            'url': 'https://mygobricks.com/',
            'source_type': 'supplier-reference',
            'notes': '公开页面说明支持自定义表、Rebrickable CSV、Studio CSV/LDR 导入，并宣称 95%+ 匹配率。',
        },
        {
            'id': 'wobrick-upload-tool',
            'name': 'Wobrick Gobricks Upload Tool',
            'domain': 'wobrick.com',
            'url': 'https://wobrick.com/toolkit',
            'source_type': 'supplier-reference',
            'notes': '公开工具页给出了 Studio CSV、BrickLink XML、Rebrickable CSV、自定义表格的导入格式与字段说明。',
        },
    ],
}

DEFAULT_JOBS: Dict[str, Any] = {
    'items': [],
}

DEFAULT_ANALYTICS: Dict[str, Any] = {
    'updated_at': '',
    'totals': {},
    'daily': {},
    'recent': [],
}


class PartAdapterStore:
    def __init__(self) -> None:
        self.lock = threading.Lock()
        self._rules_cache: Optional[Dict[str, Any]] = None
        self._rules_cache_mtime_ns: int = -1
        self._rules_summary_cache: Optional[Dict[str, Any]] = None
        self._rules_summary_cache_mtime_ns: int = -1
        self._gobricks_color_reference_cache: Optional[List[Dict[str, Any]]] = None
        self._gobricks_color_reference_cache_mtime_ns: int = -1
        PART_ADAPTER_DIR.mkdir(parents=True, exist_ok=True)
        self._ensure_seed_files()

    def _ensure_seed_files(self) -> None:
        if not RULES_PATH.exists():
            seed = deepcopy(DEFAULT_RULES)
            seed['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, seed)
        if not JOBS_PATH.exists():
            self._write_json(JOBS_PATH, deepcopy(DEFAULT_JOBS))
        if not ANALYTICS_PATH.exists():
            self._write_json(ANALYTICS_PATH, deepcopy(DEFAULT_ANALYTICS))

    def _read_json(self, path: Path, fallback: Dict[str, Any]) -> Dict[str, Any]:
        if not path.exists():
            return deepcopy(fallback)
        try:
            return json.loads(path.read_text(encoding='utf-8'))
        except Exception:
            return deepcopy(fallback)

    def _write_json(self, path: Path, payload: Dict[str, Any]) -> None:
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
        if path == RULES_PATH:
            try:
                self._rules_cache_mtime_ns = path.stat().st_mtime_ns
            except Exception:
                self._rules_cache_mtime_ns = -1
            self._rules_cache = deepcopy(payload)
            self._rules_summary_cache = None
            self._rules_summary_cache_mtime_ns = -1
            self._gobricks_color_reference_cache = None
            self._gobricks_color_reference_cache_mtime_ns = -1

    def _merge_rules_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        merged = deepcopy(DEFAULT_RULES)
        for key, value in payload.items():
            if key == 'gobricks_sync_meta' and isinstance(value, dict):
                base = merged.get('gobricks_sync_meta')
                if isinstance(base, dict):
                    base.update(value)
                else:
                    merged[key] = value
                continue
            if key in {'lego_color_catalog', 'gobricks_color_catalog', 'part_alias_map', 'print_part_map', 'external_part_map'} and isinstance(value, dict):
                base = merged.get(key)
                if isinstance(base, dict):
                    base.update(value)
                else:
                    merged[key] = value
                continue
            merged[key] = value
        lego_color_catalog = merged.get('lego_color_catalog')
        if isinstance(lego_color_catalog, dict):
            for color_id, raw_meta in list(lego_color_catalog.items()):
                meta = raw_meta if isinstance(raw_meta, dict) else {}
                name_en = str(meta.get('name_en') or '').strip()
                translated = self._translate_rebrickable_color_name(name_en, str(meta.get('name') or '').strip())
                if translated != str(meta.get('name') or '').strip():
                    updated = dict(meta)
                    updated['name'] = translated
                    lego_color_catalog[str(color_id)] = updated
        return merged

    def _get_cached_rules_payload(self) -> Dict[str, Any]:
        if not RULES_PATH.exists():
            return deepcopy(DEFAULT_RULES)
        try:
            current_mtime_ns = RULES_PATH.stat().st_mtime_ns
        except Exception:
            current_mtime_ns = -1
        if self._rules_cache is not None and current_mtime_ns == self._rules_cache_mtime_ns:
            return self._rules_cache
        payload = self._read_json(RULES_PATH, DEFAULT_RULES)
        merged = self._merge_rules_payload(payload)
        self._rules_cache = merged
        self._rules_cache_mtime_ns = current_mtime_ns
        return merged

    def _read_rules_payload(self) -> Dict[str, Any]:
        return deepcopy(self._get_cached_rules_payload())

    def get_rules(self) -> Dict[str, Any]:
        with self.lock:
            payload = deepcopy(self._get_cached_rules_payload())
        return deepcopy(payload)

    def _current_rules_mtime_ns(self) -> int:
        try:
            return RULES_PATH.stat().st_mtime_ns
        except Exception:
            return self._rules_cache_mtime_ns

    def get_rules_summary(self) -> Dict[str, Any]:
        with self.lock:
            current_mtime_ns = self._current_rules_mtime_ns()
            if self._rules_summary_cache is not None and current_mtime_ns == self._rules_summary_cache_mtime_ns:
                return deepcopy(self._rules_summary_cache)
            rules = self._get_cached_rules_payload()
            gobricks_item_index = rules.get('gobricks_item_index') if isinstance(rules.get('gobricks_item_index'), dict) else {}
            gobricks_category_index = rules.get('gobricks_category_index') if isinstance(rules.get('gobricks_category_index'), dict) else {}
            gobricks_color_catalog = rules.get('gobricks_color_catalog') if isinstance(rules.get('gobricks_color_catalog'), dict) else {}
            lego_color_catalog = rules.get('lego_color_catalog') if isinstance(rules.get('lego_color_catalog'), dict) else {}
            exact_combo_map = rules.get('exact_combo_map') if isinstance(rules.get('exact_combo_map'), dict) else {}
            shortage_combo_map = rules.get('shortage_combo_map') if isinstance(rules.get('shortage_combo_map'), dict) else {}
            exact_part_map = rules.get('exact_part_map') if isinstance(rules.get('exact_part_map'), dict) else {}
            part_alias_map = rules.get('part_alias_map') if isinstance(rules.get('part_alias_map'), dict) else {}
            color_rules = rules.get('color_rules') if isinstance(rules.get('color_rules'), dict) else {}
            substitutions = rules.get('substitutions') if isinstance(rules.get('substitutions'), dict) else {}
            part_meta = rules.get('part_meta') if isinstance(rules.get('part_meta'), dict) else {}
            part_image_cache = rules.get('part_image_cache') if isinstance(rules.get('part_image_cache'), dict) else {}
            rebrickable_part_catalog = rules.get('rebrickable_part_catalog') if isinstance(rules.get('rebrickable_part_catalog'), dict) else {}
            rebrickable_part_category_catalog = (
                rules.get('rebrickable_part_category_catalog')
                if isinstance(rules.get('rebrickable_part_category_catalog'), dict)
                else {}
            )
            element_part_map = rules.get('element_part_map') if isinstance(rules.get('element_part_map'), dict) else {}
            print_part_map = rules.get('print_part_map') if isinstance(rules.get('print_part_map'), dict) else {}
            external_part_map = rules.get('external_part_map') if isinstance(rules.get('external_part_map'), dict) else {}
            summary = {
                'updated_at': str(rules.get('updated_at') or ''),
                'gobricks_sync_meta': deepcopy(rules.get('gobricks_sync_meta') if isinstance(rules.get('gobricks_sync_meta'), dict) else {}),
                'gobricks_item_index_size': len(gobricks_item_index),
                'gobricks_category_index_size': len(gobricks_category_index),
                'gobricks_color_catalog_size': len(gobricks_color_catalog),
                'gobricks_color_catalog': deepcopy(gobricks_color_catalog),
                'part_image_cache_size': len(part_image_cache),
                'rebrickable_part_catalog_size': len(rebrickable_part_catalog),
                'rebrickable_part_category_catalog_size': len(rebrickable_part_category_catalog),
                'element_part_map_size': len(element_part_map),
                'print_part_map_size': len(print_part_map),
                'external_part_map_size': len(external_part_map),
                'lego_color_catalog_size': len(lego_color_catalog),
                'lego_color_catalog': deepcopy(lego_color_catalog),
                'exact_combo_map_size': len(exact_combo_map),
                'shortage_combo_map_size': len(shortage_combo_map),
                'exact_part_map': deepcopy(exact_part_map),
                'part_alias_map_size': len(part_alias_map),
                'color_rules': deepcopy(color_rules),
                'substitutions': deepcopy(substitutions),
                'part_meta': deepcopy(part_meta),
                'source_refs': deepcopy(rules.get('source_refs') if isinstance(rules.get('source_refs'), list) else []),
                'gobricks_color_reference': self._get_cached_gobricks_color_reference(rules, current_mtime_ns),
            }
            self._rules_summary_cache = deepcopy(summary)
            self._rules_summary_cache_mtime_ns = current_mtime_ns
            return deepcopy(summary)

    def get_sources(self) -> List[Dict[str, Any]]:
        rules = self.get_rules()
        items = rules.get('source_refs')
        return deepcopy(items if isinstance(items, list) else [])

    def _read_analytics_payload(self) -> Dict[str, Any]:
        payload = self._read_json(ANALYTICS_PATH, DEFAULT_ANALYTICS)
        merged = deepcopy(DEFAULT_ANALYTICS)
        if isinstance(payload.get('updated_at'), str):
            merged['updated_at'] = payload.get('updated_at') or ''
        if isinstance(payload.get('totals'), dict):
            merged['totals'] = deepcopy(payload.get('totals') or {})
        if isinstance(payload.get('daily'), dict):
            merged['daily'] = deepcopy(payload.get('daily') or {})
        if isinstance(payload.get('recent'), list):
            merged['recent'] = deepcopy(payload.get('recent') or [])
        return merged

    def record_event(
        self,
        event_type: str,
        route: str = '',
        source_name: str = '',
        visitor_key: str = '',
    ) -> None:
        safe_event = str(event_type or '').strip()
        if not safe_event:
            return
        with self.lock:
            payload = self._read_analytics_payload()
            today = datetime.now().strftime('%Y-%m-%d')
            daily = payload.setdefault('daily', {})
            day_bucket = daily.get(today)
            if not isinstance(day_bucket, dict):
                day_bucket = {'counts': {}, 'visitors': []}
            counts = day_bucket.get('counts') if isinstance(day_bucket.get('counts'), dict) else {}
            counts[safe_event] = int(counts.get(safe_event) or 0) + 1
            day_bucket['counts'] = counts
            visitors = day_bucket.get('visitors') if isinstance(day_bucket.get('visitors'), list) else []
            safe_visitor = str(visitor_key or '').strip()
            if safe_visitor and safe_visitor not in visitors:
                visitors.append(safe_visitor)
            day_bucket['visitors'] = visitors[-2000:]
            daily[today] = day_bucket

            totals = payload.setdefault('totals', {})
            totals[safe_event] = int(totals.get(safe_event) or 0) + 1

            recent = payload.setdefault('recent', [])
            recent.insert(0, {
                'at': _now_iso(),
                'event_type': safe_event,
                'route': str(route or '').strip(),
                'source_name': str(source_name or '').strip(),
            })
            payload['recent'] = recent[:20]

            keep_days = sorted(daily.keys())[-35:]
            payload['daily'] = {key: daily.get(key) for key in keep_days if key in daily}
            payload['updated_at'] = _now_iso()
            self._write_json(ANALYTICS_PATH, payload)

    def get_analytics_summary(self) -> Dict[str, Any]:
        with self.lock:
            payload = self._read_analytics_payload()
        daily = payload.get('daily') if isinstance(payload.get('daily'), dict) else {}
        today = datetime.now().strftime('%Y-%m-%d')
        yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        day_keys = sorted(daily.keys())
        last_7_keys = day_keys[-7:]

        def _count_for(keys: List[str], event_name: str) -> int:
            total = 0
            for key in keys:
                item = daily.get(key) if isinstance(daily.get(key), dict) else {}
                counts = item.get('counts') if isinstance(item.get('counts'), dict) else {}
                total += int(counts.get(event_name) or 0)
            return total

        def _single_day_summary(day_key: str) -> Dict[str, int]:
            bucket = daily.get(day_key) if isinstance(daily.get(day_key), dict) else {}
            counts = bucket.get('counts') if isinstance(bucket.get('counts'), dict) else {}
            visitors = bucket.get('visitors') if isinstance(bucket.get('visitors'), list) else []
            return {
                'public_home_views': int(counts.get('page_view_public') or 0),
                'public_results_views': int(counts.get('page_view_public_results') or 0),
                'admin_views': int(counts.get('page_view_admin') or 0),
                'page_views': int(counts.get('page_view_public') or 0) + int(counts.get('page_view_admin') or 0),
                'unique_visitors': len(visitors),
                'import_bom': int(counts.get('import_bom') or 0),
                'convert_gobricks': int(counts.get('convert_gobricks') or 0),
                'analyze': int(counts.get('analyze') or 0),
                'exports': int(counts.get('export_csv') or 0) + int(counts.get('export_designer_pdf') or 0) + int(counts.get('export_designer_csv') or 0),
            }

        def _funnel(summary: Dict[str, int]) -> Dict[str, Any]:
            home_views = int(summary.get('public_home_views') or 0)
            results_views = int(summary.get('public_results_views') or 0)
            uses = (
                int(summary.get('import_bom') or 0) +
                int(summary.get('convert_gobricks') or 0) +
                int(summary.get('analyze') or 0) +
                int(summary.get('exports') or 0)
            )

            def _rate(numerator: int, denominator: int) -> float:
                if denominator <= 0:
                    return 0.0
                return round((numerator / denominator) * 100, 1)

            return {
                'home_views': home_views,
                'results_views': results_views,
                'uses': uses,
                'result_view_rate': _rate(results_views, home_views),
                'use_rate': _rate(uses, home_views),
            }

        today_summary = _single_day_summary(today)
        yesterday_summary = _single_day_summary(yesterday)
        recent = payload.get('recent') if isinstance(payload.get('recent'), list) else []

        return {
            'updated_at': str(payload.get('updated_at') or ''),
            'today': today_summary,
            'yesterday': yesterday_summary,
            'last_7_days': {
                'public_home_views': _count_for(last_7_keys, 'page_view_public'),
                'public_results_views': _count_for(last_7_keys, 'page_view_public_results'),
                'admin_views': _count_for(last_7_keys, 'page_view_admin'),
                'page_views': _count_for(last_7_keys, 'page_view_public') + _count_for(last_7_keys, 'page_view_admin'),
                'import_bom': _count_for(last_7_keys, 'import_bom'),
                'convert_gobricks': _count_for(last_7_keys, 'convert_gobricks'),
                'analyze': _count_for(last_7_keys, 'analyze'),
                'exports': _count_for(last_7_keys, 'export_csv') + _count_for(last_7_keys, 'export_designer_pdf') + _count_for(last_7_keys, 'export_designer_csv'),
            },
            'funnel_today': _funnel(today_summary),
            'funnel_last_7_days': _funnel({
                'public_home_views': _count_for(last_7_keys, 'page_view_public'),
                'public_results_views': _count_for(last_7_keys, 'page_view_public_results'),
                'import_bom': _count_for(last_7_keys, 'import_bom'),
                'convert_gobricks': _count_for(last_7_keys, 'convert_gobricks'),
                'analyze': _count_for(last_7_keys, 'analyze'),
                'exports': _count_for(last_7_keys, 'export_csv') + _count_for(last_7_keys, 'export_designer_pdf') + _count_for(last_7_keys, 'export_designer_csv'),
            }),
            'totals': deepcopy(payload.get('totals') if isinstance(payload.get('totals'), dict) else {}),
            'recent': deepcopy(recent[:12]),
        }

    def _meta_to_swatch_hex(self, meta: Dict[str, Any]) -> str:
        safe = meta if isinstance(meta, dict) else {}
        family = str(safe.get('family') or 'neutral').strip().lower()
        finish = str(safe.get('finish') or 'solid').strip().lower()
        base_map: Dict[str, Any] = {
            'neutral': (148, 150, 154),
            'gray': (142, 148, 154),
            'white': (235, 237, 240),
            'black': (42, 44, 48),
            'red': (182, 58, 56),
            'yellow': (218, 188, 68),
            'blue': (74, 124, 188),
            'green': (82, 152, 92),
            'brown': (126, 84, 58),
            'tan': (190, 164, 118),
            'orange': (202, 126, 58),
            'purple': (130, 98, 170),
        }
        red, green, blue = base_map.get(family, base_map['neutral'])
        try:
            brightness = max(0, min(100, int(safe.get('brightness', 60))))
        except Exception:
            brightness = 60
        try:
            warmth = max(0, min(100, int(safe.get('warmth', 50))))
        except Exception:
            warmth = 50
        balance = (brightness - 50) / 50.0
        if balance >= 0:
            red = int(red + (255 - red) * balance * 0.65)
            green = int(green + (255 - green) * balance * 0.65)
            blue = int(blue + (255 - blue) * balance * 0.65)
        else:
            ratio = 1 + balance * 0.7
            red = int(red * ratio)
            green = int(green * ratio)
            blue = int(blue * ratio)
        warm_shift = (warmth - 50) / 50.0
        red = max(0, min(255, int(red + 18 * warm_shift)))
        blue = max(0, min(255, int(blue - 18 * warm_shift)))
        green = max(0, min(255, green))
        if finish == 'transparent':
            red = int(red + (255 - red) * 0.24)
            green = int(green + (255 - green) * 0.24)
            blue = int(blue + (255 - blue) * 0.24)
        elif finish == 'metallic':
            red = int(red * 0.9 + 24)
            green = int(green * 0.9 + 24)
            blue = int(blue * 0.9 + 24)
        elif finish == 'pearlescent':
            red = int(red + (255 - red) * 0.14)
            green = int(green + (255 - green) * 0.14)
            blue = int(blue + (255 - blue) * 0.14)
        return f'#{red:02x}{green:02x}{blue:02x}'

    def _bricklink_part_image_url(self, part_no: str, color_no: str, exact: bool = True) -> str:
        safe_part = str(part_no or '').strip()
        safe_color = str(color_no or '').strip()
        if not safe_part:
            return ''
        encoded_part = quote(safe_part, safe='')
        if exact and safe_color:
            return f'https://img.bricklink.com/ItemImage/PN/{quote(safe_color, safe="")}/{encoded_part}.png'
        return f'https://img.bricklink.com/ItemImage/PL/{encoded_part}.png'

    def _ordered_unique_strings(self, values: List[str]) -> List[str]:
        seen = set()
        result: List[str] = []
        for raw in values:
            value = str(raw or '').strip()
            if not value or value in seen:
                continue
            seen.add(value)
            result.append(value)
        return result

    def _build_gobricks_color_reference(self, rules: Dict[str, Any]) -> List[Dict[str, Any]]:
        exact_combo_map = rules.get('exact_combo_map') if isinstance(rules.get('exact_combo_map'), dict) else {}
        lego_color_catalog = rules.get('lego_color_catalog') if isinstance(rules.get('lego_color_catalog'), dict) else {}
        gobricks_color_catalog = (
            rules.get('gobricks_color_catalog') if isinstance(rules.get('gobricks_color_catalog'), dict) else {}
        )
        grouped: Dict[str, Dict[str, Any]] = {}
        for raw in exact_combo_map.values():
            item = raw if isinstance(raw, dict) else {}
            gobricks_color_no = str(item.get('gobricks_color_no') or '').strip()
            lego_color_no = str(item.get('lego_color_no') or '').strip()
            if not gobricks_color_no:
                continue
            bucket = grouped.setdefault(
                gobricks_color_no,
                {
                    'gobricks_color_no': gobricks_color_no,
                    'mapped_count': 0,
                    'lego_colors': [],
                    'finish': '',
                    'swatch_hex': '',
                    'gobricks_color_name': '',
                    'gobricks_color_name_en': '',
                    'name_source': '',
                },
            )
            bucket['mapped_count'] = int(bucket.get('mapped_count') or 0) + 1
            explicit_gobricks_meta = (
                gobricks_color_catalog.get(gobricks_color_no)
                if isinstance(gobricks_color_catalog.get(gobricks_color_no), dict)
                else {}
            )
            if explicit_gobricks_meta:
                if not bucket.get('gobricks_color_name'):
                    bucket['gobricks_color_name'] = str(explicit_gobricks_meta.get('name') or '').strip()
                if not bucket.get('gobricks_color_name_en'):
                    bucket['gobricks_color_name_en'] = str(explicit_gobricks_meta.get('name_en') or '').strip()
                if not bucket.get('swatch_hex') and str(explicit_gobricks_meta.get('swatch_hex') or '').strip():
                    bucket['swatch_hex'] = str(explicit_gobricks_meta.get('swatch_hex') or '').strip()
                if not bucket.get('name_source') and (
                    bucket.get('gobricks_color_name') or bucket.get('gobricks_color_name_en')
                ):
                    bucket['name_source'] = 'gobricks_catalog'
            if lego_color_no and lego_color_no not in {entry['lego_color_no'] for entry in bucket['lego_colors']}:
                meta = lego_color_catalog.get(lego_color_no) if isinstance(lego_color_catalog.get(lego_color_no), dict) else {}
                bucket['lego_colors'].append(
                    {
                        'lego_color_no': lego_color_no,
                        'lego_color_name': str(meta.get('name') or '').strip(),
                        'lego_color_name_en': str(meta.get('name_en') or '').strip(),
                    }
                )
                if not bucket.get('finish') and str(meta.get('finish') or '').strip():
                    bucket['finish'] = str(meta.get('finish') or '').strip()
                if not bucket.get('swatch_hex'):
                    bucket['swatch_hex'] = self._meta_to_swatch_hex(meta)
                if not bucket.get('gobricks_color_name'):
                    bucket['gobricks_color_name'] = str(meta.get('name') or '').strip()
                    bucket['name_source'] = 'mapped_inference'
                if not bucket.get('gobricks_color_name_en'):
                    bucket['gobricks_color_name_en'] = str(meta.get('name_en') or '').strip()

        def sort_key(item: Dict[str, Any]) -> Any:
            raw = str(item.get('gobricks_color_no') or '')
            try:
                return (0, int(raw))
            except Exception:
                return (1, raw)

        items = sorted(grouped.values(), key=sort_key)
        result = []
        for row in items:
            lego_colors = row.get('lego_colors') if isinstance(row.get('lego_colors'), list) else []
            lego_labels = []
            for entry in lego_colors[:5]:
                safe = entry if isinstance(entry, dict) else {}
                lego_labels.append(
                    {
                        'lego_color_no': str(safe.get('lego_color_no') or '').strip(),
                        'lego_color_name': str(safe.get('lego_color_name') or '').strip(),
                        'lego_color_name_en': str(safe.get('lego_color_name_en') or '').strip(),
                    }
                )
            result.append(
                {
                    'gobricks_color_no': str(row.get('gobricks_color_no') or '').strip(),
                    'gobricks_color_name': str(row.get('gobricks_color_name') or '').strip(),
                    'gobricks_color_name_en': str(row.get('gobricks_color_name_en') or '').strip(),
                    'name_source': str(row.get('name_source') or '').strip(),
                    'mapped_count': int(row.get('mapped_count') or 0),
                    'finish': str(row.get('finish') or '').strip(),
                    'swatch_hex': str(row.get('swatch_hex') or '').strip(),
                    'lego_colors': lego_labels,
                }
            )
        return result

    def get_gobricks_color_reference(self) -> List[Dict[str, Any]]:
        with self.lock:
            rules = self._get_cached_rules_payload()
            return deepcopy(self._get_cached_gobricks_color_reference(rules, self._current_rules_mtime_ns()))

    def _get_cached_gobricks_color_reference(self, rules: Dict[str, Any], mtime_ns: int) -> List[Dict[str, Any]]:
        if self._gobricks_color_reference_cache is not None and mtime_ns == self._gobricks_color_reference_cache_mtime_ns:
            return self._gobricks_color_reference_cache
        items = self._build_gobricks_color_reference(rules)
        self._gobricks_color_reference_cache = items
        self._gobricks_color_reference_cache_mtime_ns = mtime_ns
        return items

    def update_rules(
        self,
        gobricks_sync_meta: Optional[Dict[str, Any]] = None,
        gobricks_item_index: Optional[Dict[str, Any]] = None,
        gobricks_category_index: Optional[Dict[str, Any]] = None,
        exact_combo_map: Optional[Dict[str, Any]] = None,
        shortage_combo_map: Optional[Dict[str, Any]] = None,
        lego_color_catalog: Optional[Dict[str, Any]] = None,
        exact_part_map: Optional[Dict[str, Any]] = None,
        part_alias_map: Optional[Dict[str, Any]] = None,
        color_rules: Optional[Dict[str, Any]] = None,
        substitutions: Optional[Dict[str, Any]] = None,
        part_meta: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        with self.lock:
            payload = self._read_rules_payload()
            if gobricks_sync_meta is not None:
                payload['gobricks_sync_meta'] = self._normalize_keyed_dict(gobricks_sync_meta)
            if gobricks_item_index is not None:
                payload['gobricks_item_index'] = self._normalize_keyed_dict(gobricks_item_index)
            if gobricks_category_index is not None:
                payload['gobricks_category_index'] = self._normalize_keyed_dict(gobricks_category_index)
            if exact_combo_map is not None:
                payload['exact_combo_map'] = self._normalize_keyed_dict(exact_combo_map)
            if shortage_combo_map is not None:
                payload['shortage_combo_map'] = self._normalize_keyed_dict(shortage_combo_map)
            if lego_color_catalog is not None:
                payload['lego_color_catalog'] = self._normalize_keyed_dict(lego_color_catalog)
            if exact_part_map is not None:
                payload['exact_part_map'] = self._normalize_keyed_dict(exact_part_map)
            if part_alias_map is not None:
                payload['part_alias_map'] = self._normalize_keyed_dict(part_alias_map)
            if color_rules is not None:
                payload['color_rules'] = self._normalize_keyed_dict(color_rules)
            if substitutions is not None:
                payload['substitutions'] = self._normalize_keyed_dict(substitutions)
            if part_meta is not None:
                payload['part_meta'] = self._normalize_keyed_dict(part_meta)
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)
            return deepcopy(payload)

    def _normalize_keyed_dict(self, value: Any) -> Dict[str, Any]:
        if not isinstance(value, dict):
            return {}
        result: Dict[str, Any] = {}
        for raw_key, raw_value in value.items():
            key = str(raw_key or '').strip()
            if not key:
                continue
            result[key] = raw_value
        return result

    def _safe_number(self, value: Any, integer: bool) -> Any:
        raw = str(value or '').strip()
        if not raw:
            return 0 if integer else 0.0
        try:
            return int(float(raw)) if integer else float(raw)
        except Exception:
            return 0 if integer else 0.0

    def _infer_color_family(self, name: str) -> str:
        text = str(name or '').strip().lower()
        if not text:
            return 'neutral'
        if any(token in text for token in ('black', 'gray', 'grey', 'silver')):
            return 'gray'
        if any(token in text for token in ('white', 'clear', 'transparent')):
            return 'white'
        if any(token in text for token in ('red', 'pink', 'magenta')):
            return 'red'
        if any(token in text for token in ('yellow', 'gold')):
            return 'yellow'
        if any(token in text for token in ('blue', 'azure', 'teal', 'turquoise')):
            return 'blue'
        if any(token in text for token in ('green', 'lime', 'olive')):
            return 'green'
        if any(token in text for token in ('brown', 'tan', 'nougat', 'beige')):
            return 'brown'
        if 'orange' in text:
            return 'orange'
        if any(token in text for token in ('purple', 'lavender', 'violet', 'lilac')):
            return 'purple'
        return 'neutral'

    def _rgb_to_tuple(self, rgb: str) -> Any:
        value = str(rgb or '').strip().lstrip('#')
        if len(value) != 6:
            return (None, None, None)
        try:
            return int(value[0:2], 16), int(value[2:4], 16), int(value[4:6], 16)
        except Exception:
            return (None, None, None)

    def _estimate_brightness(self, red: Any, green: Any, blue: Any) -> Optional[int]:
        if red is None or green is None or blue is None:
            return None
        value = round(((0.299 * red) + (0.587 * green) + (0.114 * blue)) / 255 * 100)
        return max(0, min(100, value))

    def _estimate_warmth(self, red: Any, blue: Any) -> Optional[int]:
        if red is None or blue is None:
            return None
        value = round(((red - blue) + 255) / 510 * 100)
        return max(0, min(100, value))

    def _translate_rebrickable_color_name(self, name_en: str, fallback_zh: str = '') -> str:
        safe_en = str(name_en or '').strip()
        if not safe_en:
            return str(fallback_zh or '').strip()
        mapped = REBRICKABLE_COLOR_NAME_ZH.get(safe_en)
        if mapped:
            return mapped
        return str(fallback_zh or '').strip()

    def _get_rebrickable_api_key(self) -> str:
        return os.getenv('REBRICKABLE_API_KEY', '').strip()

    def _flatten_external_ids(self, raw_external: Any) -> Dict[str, List[str]]:
        if not isinstance(raw_external, dict):
            return {}
        result: Dict[str, List[str]] = {}
        for raw_key, raw_value in raw_external.items():
            key = str(raw_key or '').strip()
            if not key:
                continue
            values: List[str] = []
            if isinstance(raw_value, list):
                values = [str(item or '').strip() for item in raw_value if str(item or '').strip()]
            elif isinstance(raw_value, dict):
                for nested in raw_value.values():
                    if isinstance(nested, list):
                        values.extend(str(item or '').strip() for item in nested if str(item or '').strip())
                    else:
                        nested_text = str(nested or '').strip()
                        if nested_text:
                            values.append(nested_text)
            else:
                text = str(raw_value or '').strip()
                if text:
                    values = [text]
            if values:
                result[key] = values
        return result

    def _fetch_rebrickable_part_details(self, part_nos: List[str]) -> Dict[str, Dict[str, Any]]:
        api_key = self._get_rebrickable_api_key()
        normalized = []
        seen = set()
        for raw in part_nos:
            part_no = str(raw or '').strip()
            if not part_no or part_no in seen:
                continue
            seen.add(part_no)
            normalized.append(part_no)
        if not api_key or not normalized:
            return {}

        result: Dict[str, Dict[str, Any]] = {}
        chunk_size = 20
        for index in range(0, len(normalized), chunk_size):
            chunk = normalized[index:index + chunk_size]
            query = quote(','.join(chunk), safe=',')
            url = f'https://rebrickable.com/api/v3/lego/parts/?part_nums={query}&inc_part_details=1&page_size={len(chunk)}'
            req = Request(url, headers={'Authorization': f'key {api_key}', 'User-Agent': 'Mozilla/5.0'})
            try:
                with urlopen(req, timeout=20) as resp:
                    body = resp.read().decode('utf-8', errors='replace')
            except HTTPError:
                continue
            except Exception:
                continue
            try:
                payload = json.loads(body or '{}')
            except Exception:
                continue
            items = payload.get('results') if isinstance(payload.get('results'), list) else []
            for raw_item in items:
                item = raw_item if isinstance(raw_item, dict) else {}
                part_num = str(item.get('part_num') or '').strip()
                if not part_num:
                    continue
                image_url = str(
                    item.get('part_img_url')
                    or item.get('part_img_url_full')
                    or item.get('element_img_url')
                    or ''
                ).strip()
                part_name = str(item.get('name') or '').strip()
                external_ids = self._flatten_external_ids(item.get('external_ids'))
                if image_url or part_name or external_ids:
                    result[part_num] = {
                        'image_url': image_url,
                        'name': part_name,
                        'external_ids': external_ids,
                    }
        return result

    def _ensure_part_details(self, part_nos: List[str], rules: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
        detail_cache = rules.get('part_detail_cache') if isinstance(rules.get('part_detail_cache'), dict) else {}
        cache = rules.get('part_image_cache') if isinstance(rules.get('part_image_cache'), dict) else {}
        normalized = []
        for raw in part_nos:
            part_no = str(raw or '').strip()
            if not part_no:
                continue
            normalized.append(part_no)
        unique_part_nos = sorted(set(normalized))
        if not unique_part_nos:
            return {}

        found: Dict[str, Dict[str, Any]] = {}
        for part_no in unique_part_nos:
            cached_detail = detail_cache.get(part_no) if isinstance(detail_cache.get(part_no), dict) else {}
            cached_image = str(cache.get(part_no) or '').strip()
            image_url = str(cached_detail.get('image_url') or cached_image).strip()
            part_name = str(cached_detail.get('name') or '').strip()
            external_ids = self._flatten_external_ids(cached_detail.get('external_ids'))
            if image_url or part_name or external_ids:
                found[part_no] = {'image_url': image_url, 'name': part_name, 'external_ids': external_ids}
        missing = [part_no for part_no in unique_part_nos if part_no not in found]
        if not missing:
            return found

        fetched = self._fetch_rebrickable_part_details(missing)
        if not fetched:
            return found

        with self.lock:
            payload = self._read_rules_payload()
            live_detail_cache = payload.get('part_detail_cache') if isinstance(payload.get('part_detail_cache'), dict) else {}
            live_cache = payload.get('part_image_cache') if isinstance(payload.get('part_image_cache'), dict) else {}
            external_part_map = payload.get('external_part_map') if isinstance(payload.get('external_part_map'), dict) else {}
            for key, value in fetched.items():
                safe_key = str(key or '').strip()
                detail = value if isinstance(value, dict) else {}
                image_url = str(detail.get('image_url') or '').strip()
                part_name = str(detail.get('name') or '').strip()
                external_ids = self._flatten_external_ids(detail.get('external_ids'))
                if image_url:
                    live_cache[safe_key] = image_url
                if image_url or part_name or external_ids:
                    live_detail_cache[safe_key] = {
                        'image_url': image_url,
                        'name': part_name,
                        'external_ids': external_ids,
                    }
                for system, values in external_ids.items():
                    for raw_value in values:
                        normalized = str(raw_value or '').strip().lower()
                        if not normalized:
                            continue
                        system_key = str(system or '').strip().lower()
                        external_part_map[f'{system_key}:{normalized}'] = safe_key
                        if system_key == 'ldraw' and normalized.endswith('.dat'):
                            external_part_map[f'{system_key}:{normalized[:-4]}'] = safe_key
                        if system_key == 'bricklink':
                            external_part_map[f'part:{normalized}'] = safe_key
            payload['external_part_map'] = external_part_map
            payload['part_detail_cache'] = live_detail_cache
            payload['part_image_cache'] = live_cache
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)

        found.update(fetched)
        return found

    def list_jobs(self, limit: int = 20) -> Dict[str, Any]:
        safe_limit = max(1, min(int(limit or 20), 200))
        with self.lock:
            payload = self._read_json(JOBS_PATH, DEFAULT_JOBS)
        items = payload.get('items')
        source = items if isinstance(items, list) else []
        slim = []
        for item in source[:safe_limit]:
            summary = item.get('summary') if isinstance(item, dict) else {}
            project = item.get('project') if isinstance(item, dict) else {}
            slim.append(
                {
                    'job_id': item.get('job_id'),
                    'project_name': project.get('project_name', ''),
                    'designer_name': project.get('designer_name', ''),
                    'source_name': project.get('source_name', ''),
                    'created_at': item.get('created_at', ''),
                    'auto_match_rate': summary.get('auto_match_rate', 0),
                    'review_qty': summary.get('review_qty', 0),
                    'blocked_qty': summary.get('blocked_qty', 0),
                    'total_qty': summary.get('total_qty', 0),
                }
            )
        return {'items': slim, 'total': len(source)}

    def get_job(self, job_id: str) -> Optional[Dict[str, Any]]:
        safe_job_id = str(job_id or '').strip()
        if not safe_job_id:
            return None
        with self.lock:
            payload = self._read_json(JOBS_PATH, DEFAULT_JOBS)
        items = payload.get('items')
        source = items if isinstance(items, list) else []
        for item in source:
            if str(item.get('job_id') or '') == safe_job_id:
                return deepcopy(item)
        return None

    def analyze(
        self,
        project_name: str,
        designer_name: str,
        source_name: str,
        bom_text: str,
        color_mode: str = 'balanced',
        optimizer_mode: str = 'reliability',
        optimizer_profile: str = 'v1',
        allow_display_sub: bool = True,
        allow_structural_sub: bool = False,
    ) -> Dict[str, Any]:
        overall_started = time.perf_counter()
        timing_marks: Dict[str, float] = {}

        def mark(name: str, started_at: float) -> None:
            timing_marks[name] = round((time.perf_counter() - started_at) * 1000, 2)

        step_started = time.perf_counter()
        with self.lock:
            rules = self._get_cached_rules_payload()
        mark('load_rules_ms', step_started)

        step_started = time.perf_counter()
        analysis_cache = self._build_analysis_cache(rules)
        mark('build_indexes_ms', step_started)

        step_started = time.perf_counter()
        parsed = self._parse_bom_text(bom_text)
        mark('parse_bom_ms', step_started)

        step_started = time.perf_counter()
        resolve_cache: Dict[str, Dict[str, Any]] = {}
        cache_hits = 0
        cache_misses = 0
        results: List[Dict[str, Any]] = []
        for row in parsed:
            cache_key = '||'.join(
                [
                    str(row.get('part_no') or '').strip(),
                    str(row.get('color_no') or '').strip(),
                    str(row.get('qty') or 1),
                    str(row.get('tag') or '').strip().lower(),
                    str(row.get('base_part_no') or '').strip(),
                    str(color_mode or '').strip().lower(),
                    str(optimizer_mode or '').strip().lower(),
                    '1' if allow_display_sub else '0',
                    '1' if allow_structural_sub else '0',
                ]
            )
            cached = resolve_cache.get(cache_key)
            if isinstance(cached, dict):
                cache_hits += 1
                resolved = deepcopy(cached)
                for field in ('line_no', 'part_no', 'color_no', 'color_name', 'qty', 'name', 'tag', 'base_part_no'):
                    resolved[field] = row.get(field)
                results.append(resolved)
                continue
            cache_misses += 1
            resolved = self._resolve_row(
                item=row,
                color_mode=color_mode,
                optimizer_mode=optimizer_mode,
                allow_display_sub=allow_display_sub,
                allow_structural_sub=allow_structural_sub,
                rules=rules,
                analysis_cache=analysis_cache,
            )
            resolve_cache[cache_key] = deepcopy(resolved)
            results.append(resolved)
        mark('resolve_rows_ms', step_started)

        step_started = time.perf_counter()
        results = self._decorate_result_rows(results=results, rules=rules)
        mark('decorate_rows_ms', step_started)

        step_started = time.perf_counter()
        summary = self._build_summary(results)
        mark('build_summary_ms', step_started)
        timing_marks['row_count'] = len(parsed)
        timing_marks['result_count'] = len(results)
        timing_marks['resolve_cache_hits'] = cache_hits
        timing_marks['resolve_cache_misses'] = cache_misses
        timing_marks['resolve_cache_size'] = len(resolve_cache)
        timing_marks['avg_resolve_row_ms'] = round(
            timing_marks.get('resolve_rows_ms', 0) / max(1, len(parsed)),
            2,
        )
        timing_marks['pre_write_total_ms'] = round((time.perf_counter() - overall_started) * 1000, 2)
        timing_marks['pre_write_total_seconds'] = round(timing_marks['pre_write_total_ms'] / 1000, 2)
        summary['performance'] = deepcopy(timing_marks)
        job = {
            'job_id': f'pa-{uuid.uuid4().hex[:12]}',
            'project': {
                'project_name': str(project_name or '').strip(),
                'designer_name': str(designer_name or '').strip(),
                'source_name': str(source_name or '').strip(),
            },
            'strategy': {
                'color_mode': color_mode,
                'optimizer_mode': optimizer_mode,
                'optimizer_profile': optimizer_profile,
                'allow_display_sub': bool(allow_display_sub),
                'allow_structural_sub': bool(allow_structural_sub),
            },
            'results': results,
            'summary': summary,
            'performance': deepcopy(timing_marks),
            'created_at': _now_iso(),
            'updated_at': _now_iso(),
        }
        step_started = time.perf_counter()
        with self.lock:
            payload = self._read_json(JOBS_PATH, DEFAULT_JOBS)
            items = payload.get('items')
            source = items if isinstance(items, list) else []
            source.insert(0, job)
            payload['items'] = source[:200]
            self._write_json(JOBS_PATH, payload)
        mark('write_job_ms', step_started)
        timing_marks['total_ms'] = round((time.perf_counter() - overall_started) * 1000, 2)
        timing_marks['total_seconds'] = round(timing_marks['total_ms'] / 1000, 2)
        summary['performance'] = deepcopy(timing_marks)
        job['summary'] = summary
        job['performance'] = deepcopy(timing_marks)
        print(
            '[part-adapter.analyze]',
            json.dumps(
                {
                    'source_name': str(source_name or '').strip(),
                    'project_name': str(project_name or '').strip(),
                    'timings': timing_marks,
                },
                ensure_ascii=False,
            ),
        )
        return deepcopy(job)

    def _build_analysis_cache(self, rules: Dict[str, Any]) -> Dict[str, Any]:
        exact_combo_map = rules.get('exact_combo_map') if isinstance(rules.get('exact_combo_map'), dict) else {}
        gobricks_item_index = rules.get('gobricks_item_index') if isinstance(rules.get('gobricks_item_index'), dict) else {}
        exact_combo_part_index: Dict[str, List[Dict[str, Any]]] = {}
        gobricks_item_part_index: Dict[str, List[Dict[str, Any]]] = {}

        for raw in exact_combo_map.values():
            entry = raw if isinstance(raw, dict) else {}
            lego_part_no = str(entry.get('lego_part_no') or '').strip()
            if not lego_part_no:
                continue
            lego_base = self._extract_base_part_no(part_no=lego_part_no, ldraw_id='')
            for key in {lego_part_no, lego_base}:
                safe_key = str(key or '').strip()
                if not safe_key:
                    continue
                exact_combo_part_index.setdefault(safe_key, []).append(entry)

        for raw in gobricks_item_index.values():
            item = raw if isinstance(raw, dict) else {}
            lego_id = str(item.get('lego_id') or '').strip()
            if not lego_id:
                continue
            lego_base = self._extract_base_part_no(part_no=lego_id, ldraw_id='')
            for key in {lego_id, lego_base}:
                safe_key = str(key or '').strip()
                if not safe_key:
                    continue
                gobricks_item_part_index.setdefault(safe_key, []).append(item)

        return {
            'exact_combo_part_index': exact_combo_part_index,
            'gobricks_item_part_index': gobricks_item_part_index,
            'expand_part_candidates_cache': {},
            'color_pair_score_cache': {},
        }

    def import_gobricks_result_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        safe_name = str(filename or '').strip()
        rows = self._read_delimited_rows(content)
        if not rows:
            raise ValueError('导入文件为空或无法解析')
        header_index = self._build_header_index(rows[0])
        if 'part_no' not in header_index or 'color_no' not in header_index:
            raise ValueError('未识别到高砖结果文件所需的乐高零件列')

        imported = 0
        conversion_rows = 0
        shortage_rows = 0
        with self.lock:
            payload = self._read_rules_payload()
            exact_combo_map = payload.get('exact_combo_map') if isinstance(payload.get('exact_combo_map'), dict) else {}
            shortage_combo_map = payload.get('shortage_combo_map') if isinstance(payload.get('shortage_combo_map'), dict) else {}

            is_conversion = any('gobrick' in ''.join(str(cell or '').lower().split()) for cell in rows[0])
            for raw_row in rows[1:]:
                values = [str(cell or '').strip() for cell in raw_row]
                if not any(values):
                    continue
                part_no = self._pick_from_row(values, header_index, 'part_no')
                color_no = self._pick_from_row(values, header_index, 'color_no')
                if not part_no:
                    continue
                combo_key = self._combo_key(part_no=part_no, color_no=color_no)
                if is_conversion:
                    gobrick_part = self._pick_gobricks_part(values, rows[0])
                    gobrick_color = self._pick_gobricks_color(values, rows[0])
                    if not self._is_valid_gobricks_mapping(gobrick_part, gobrick_color):
                        continue
                    exact_combo_map[combo_key] = {
                        'lego_part_no': part_no,
                        'lego_color_no': color_no,
                        'gobricks_part_no': gobrick_part,
                        'gobricks_color_no': gobrick_color,
                        'source_file': safe_name,
                        'updated_at': _now_iso(),
                    }
                    imported += 1
                    conversion_rows += 1
                else:
                    shortage_type = self._pick_shortage_type(values, rows[0])
                    shortage_combo_map[combo_key] = {
                        'lego_part_no': part_no,
                        'lego_color_no': color_no,
                        'shortage_type': shortage_type,
                        'source_file': safe_name,
                        'updated_at': _now_iso(),
                    }
                    imported += 1
                    shortage_rows += 1

            payload['exact_combo_map'] = exact_combo_map
            payload['shortage_combo_map'] = shortage_combo_map
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)

        return {
            'filename': safe_name,
            'imported': imported,
            'conversion_rows': conversion_rows,
            'shortage_rows': shortage_rows,
            'mode': 'conversion' if is_conversion else 'shortage',
            'exact_combo_total': len(exact_combo_map),
            'shortage_combo_total': len(shortage_combo_map),
        }

    def sync_gobricks_items(
        self,
        items: List[Dict[str, Any]],
        start_time: str,
        end_time: str,
        base_url: str,
        need_detail_info: bool,
    ) -> Dict[str, Any]:
        normalized_items = items if isinstance(items, list) else []
        with self.lock:
            payload = self._read_rules_payload()
            index_map = payload.get('gobricks_item_index') if isinstance(payload.get('gobricks_item_index'), dict) else {}
            updated = 0
            for raw in normalized_items:
                item = raw if isinstance(raw, dict) else {}
                item_id = str(item.get('item_id') or '').strip()
                if not item_id:
                    continue
                prev = index_map.get(item_id) if isinstance(index_map.get(item_id), dict) else {}
                inventory = item.get('inventory', prev.get('inventory', 0))
                status_value = item.get('status', prev.get('status', 0))
                price_value = item.get('price', prev.get('price', 0))
                merged = {
                    'item_id': item_id,
                    'lego_id': str(item.get('lego_id') or prev.get('lego_id') or '').strip(),
                    'lego_color_id': str(item.get('lego_color_id') or prev.get('lego_color_id') or '').strip(),
                    'color_id': str(item.get('color_id') or prev.get('color_id') or '').strip(),
                    'inventory': inventory,
                    'status': status_value,
                    'price': price_value,
                    'caption': str(item.get('caption') or prev.get('caption') or '').strip(),
                    'caption_en': str(item.get('caption_en') or prev.get('caption_en') or '').strip(),
                    'change_time': str(item.get('change_time') or prev.get('change_time') or '').strip(),
                    'product_id': str(item.get('product_id') or prev.get('product_id') or '').strip(),
                    'product_weight': str(item.get('product_weight') or prev.get('product_weight') or '').strip(),
                    'category_path': str(item.get('category_path') or prev.get('category_path') or '').strip(),
                    'category_name': str(item.get('category_name') or prev.get('category_name') or '').strip(),
                    'created_time': str(item.get('created_time') or prev.get('created_time') or '').strip(),
                    'on_shelf_time': str(item.get('on_shelf_time') or prev.get('on_shelf_time') or '').strip(),
                    'updated_at': _now_iso(),
                }
                index_map[item_id] = merged
                updated += 1

            payload['gobricks_item_index'] = index_map
            payload['gobricks_sync_meta'] = {
                'last_sync_at': _now_iso(),
                'last_sync_count': updated,
                'last_sync_start_time': str(start_time or '').strip(),
                'last_sync_end_time': str(end_time or '').strip(),
                'base_url': str(base_url or 'https://api.gobricks.cn').strip() or 'https://api.gobricks.cn',
                'need_detail_info': bool(need_detail_info),
            }
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)

        return {
            'updated': updated,
            'total_indexed': len(index_map),
            'last_sync_at': payload['gobricks_sync_meta']['last_sync_at'],
        }

    def import_gobricks_catalog_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        safe_name = str(filename or '').strip()
        if Path(safe_name).suffix.lower() != '.xlsx':
            raise ValueError('高砖零件表仅支持 xlsx 文件')
        rows = self._read_xlsx_rows(content)
        if not rows:
            raise ValueError('高砖零件表为空')
        header = [str(cell or '').strip() for cell in rows[0]]
        header_map = {''.join(str(cell or '').strip().split()): idx for idx, cell in enumerate(header)}
        item_id_key = 'item_id' if 'item_id' in header_map else 'id' if 'id' in header_map else ''
        caption_key = 'caption' if 'caption' in header_map else '名称' if '名称' in header_map else ''
        inventory_key = 'inventory' if 'inventory' in header_map else '库存' if '库存' in header_map else ''
        price_key = 'price' if 'price' in header_map else '价格' if '价格' in header_map else ''
        status_key = 'shelf_state' if 'shelf_state' in header_map else '上下架' if '上下架' in header_map else ''
        picture_key = 'picture' if 'picture' in header_map else ''
        product_id_key = 'product_id' if 'product_id' in header_map else ''
        color_id_key = 'color_id' if 'color_id' in header_map else ''
        lego_id_key = 'lego_id' if 'lego_id' in header_map else ''
        lego_color_id_key = 'lego_color_id' if 'lego_color_id' in header_map else ''
        category_key = 'ldd_catalog' if 'ldd_catalog' in header_map else '分类' if '分类' in header_map else ''
        weight_key = 'product_weight' if 'product_weight' in header_map else '单重' if '单重' in header_map else ''
        created_key = 'created_at' if 'created_at' in header_map else '创建时间' if '创建时间' in header_map else ''
        on_shelf_key = 'change_time' if 'change_time' in header_map else '上架时间' if '上架时间' in header_map else ''

        required_labels = []
        if not item_id_key:
            required_labels.append('item_id / id')
        if not caption_key:
            required_labels.append('caption / 名称')
        if not inventory_key:
            required_labels.append('inventory / 库存')
        if not price_key:
            required_labels.append('price / 价格')
        if not status_key:
            required_labels.append('shelf_state / 上下架')
        if required_labels:
            raise ValueError(f'高砖零件表缺少必要列：{", ".join(required_labels)}')

        imported = 0
        with self.lock:
            payload = self._read_rules_payload()
            index_map = payload.get('gobricks_item_index') if isinstance(payload.get('gobricks_item_index'), dict) else {}
            for raw in rows[1:]:
                values = [str(cell).strip() if cell is not None else '' for cell in raw]
                if not any(values):
                    continue
                item_id = values[header_map[item_id_key]].strip()
                if not item_id:
                    continue
                target = self._extract_gobricks_target(item_id=item_id)
                category_path = values[header_map.get(category_key, -1)].strip() if category_key and header_map.get(category_key, -1) >= 0 else ''
                category_name = category_path.split(',')[-1].strip() if category_path else ''
                prev = index_map.get(item_id) if isinstance(index_map.get(item_id), dict) else {}
                picture_url = values[header_map.get(picture_key, -1)].strip() if picture_key and header_map.get(picture_key, -1) >= 0 else ''
                product_id = values[header_map.get(product_id_key, -1)].strip() if product_id_key and header_map.get(product_id_key, -1) >= 0 else ''
                color_id = values[header_map.get(color_id_key, -1)].strip() if color_id_key and header_map.get(color_id_key, -1) >= 0 else ''
                lego_id = values[header_map.get(lego_id_key, -1)].strip() if lego_id_key and header_map.get(lego_id_key, -1) >= 0 else ''
                lego_color_id = values[header_map.get(lego_color_id_key, -1)].strip() if lego_color_id_key and header_map.get(lego_color_id_key, -1) >= 0 else ''
                index_map[item_id] = {
                    'item_id': item_id,
                    'lego_id': lego_id or str(prev.get('lego_id') or '').strip(),
                    'lego_color_id': lego_color_id or str(prev.get('lego_color_id') or '').strip(),
                    'color_id': color_id or str(target.get('color_no') or prev.get('color_id') or '').strip(),
                    'inventory': self._safe_number(values[header_map.get(inventory_key, -1)] if inventory_key and header_map.get(inventory_key, -1) >= 0 else '', integer=True),
                    'status': values[header_map.get(status_key, -1)].strip() if status_key and header_map.get(status_key, -1) >= 0 else str(prev.get('status') or ''),
                    'price': self._safe_number(values[header_map.get(price_key, -1)] if price_key and header_map.get(price_key, -1) >= 0 else '', integer=False),
                    'caption': values[header_map.get(caption_key, -1)].strip() if caption_key and header_map.get(caption_key, -1) >= 0 else str(prev.get('caption') or ''),
                    'caption_en': str(prev.get('caption_en') or '').strip(),
                    'change_time': str(prev.get('change_time') or '').strip(),
                    'product_id': product_id or str(target.get('part_no') or '').replace('GDS-', '').strip() or str(prev.get('product_id') or '').strip(),
                    'product_weight': values[header_map.get(weight_key, -1)].strip() if weight_key and header_map.get(weight_key, -1) >= 0 else str(prev.get('product_weight') or ''),
                    'category_path': category_path,
                    'category_name': category_name,
                    'image_url': picture_url or str(prev.get('image_url') or '').strip(),
                    'created_time': values[header_map.get(created_key, -1)].strip() if created_key and header_map.get(created_key, -1) >= 0 else str(prev.get('created_time') or ''),
                    'on_shelf_time': values[header_map.get(on_shelf_key, -1)].strip() if on_shelf_key and header_map.get(on_shelf_key, -1) >= 0 else str(prev.get('on_shelf_time') or ''),
                    'updated_at': _now_iso(),
                    'source_file': safe_name,
                }
                imported += 1
            payload['gobricks_item_index'] = index_map
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)
        return {'filename': safe_name, 'imported': imported, 'total_indexed': len(index_map)}

    def import_rebrickable_relationships_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        safe_name = str(filename or '').strip()
        rows = self._read_delimited_rows(content)
        if not rows:
            raise ValueError('Rebrickable 关系表为空或无法解析')
        header = [str(cell or '').strip() for cell in rows[0]]
        header_map = {''.join(str(cell or '').strip().lower().split()): idx for idx, cell in enumerate(header)}

        def pick(values: List[str], keys: List[str]) -> str:
            for key in keys:
                idx = header_map.get(key)
                if idx is None:
                    continue
                if idx < len(values):
                    value = str(values[idx] or '').strip()
                    if value:
                        return value
            return ''

        required_groups = {
            'rel_type': ['rel_type', 'relationshiptype', 'reltype'],
            'child': ['child_part_num', 'childpartnum', 'child_part', 'childpart'],
            'parent': ['parent_part_num', 'parentpartnum', 'parent_part', 'parentpart'],
        }
        missing = [name for name, keys in required_groups.items() if not any(key in header_map for key in keys)]
        if missing:
            raise ValueError('未识别到 Rebrickable part_relationships.csv 所需列（需要 rel_type / child_part_num / parent_part_num）')

        imported = 0
        rel_counts = {'M': 0, 'A': 0, 'P': 0}
        with self.lock:
            payload = self._read_rules_payload()
            alias_map = payload.get('part_alias_map') if isinstance(payload.get('part_alias_map'), dict) else {}
            print_part_map = payload.get('print_part_map') if isinstance(payload.get('print_part_map'), dict) else {}
            merged = {str(k).strip(): str(v).strip() for k, v in alias_map.items() if str(k).strip() and str(v).strip()}
            print_merged = {
                str(k).strip(): str(v).strip()
                for k, v in print_part_map.items()
                if str(k).strip() and str(v).strip()
            }

            for raw in rows[1:]:
                values = [str(cell or '').strip() for cell in raw]
                if not any(values):
                    continue
                rel_type = pick(values, required_groups['rel_type']).upper()
                if rel_type not in {'M', 'A', 'P'}:
                    continue
                child = pick(values, required_groups['child'])
                parent = pick(values, required_groups['parent'])
                if not child or not parent or child == parent:
                    continue
                if rel_type == 'P':
                    print_merged[child] = parent
                else:
                    merged[child] = parent
                imported += 1
                rel_counts[rel_type] = rel_counts.get(rel_type, 0) + 1

            def resolve(value: str) -> str:
                current = str(value or '').strip()
                seen = set()
                while current and current not in seen:
                    seen.add(current)
                    nxt = str(merged.get(current) or '').strip()
                    if not nxt:
                        break
                    current = nxt
                return current or str(value or '').strip()

            flattened: Dict[str, str] = {}
            for raw_child, raw_parent in merged.items():
                child = str(raw_child or '').strip()
                parent = resolve(raw_parent)
                if child and parent and child != parent:
                    flattened[child] = parent

            payload['part_alias_map'] = flattened
            payload['print_part_map'] = print_merged
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)

        return {
            'filename': safe_name,
            'imported': imported,
            'rel_m_count': rel_counts.get('M', 0),
            'rel_a_count': rel_counts.get('A', 0),
            'rel_p_count': rel_counts.get('P', 0),
            'total_aliases': len(flattened),
            'total_print_relations': len(print_merged),
        }

    def import_rebrickable_parts_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        safe_name = str(filename or '').strip()
        rows = self._read_delimited_rows(content)
        if not rows:
            raise ValueError('Rebrickable parts.csv 为空或无法解析')
        header = [str(cell or '').strip() for cell in rows[0]]
        header_map = {''.join(str(cell or '').strip().lower().split()): idx for idx, cell in enumerate(header)}
        required = ['part_num', 'name', 'part_cat_id']
        missing = [key for key in required if key not in header_map]
        if missing:
            raise ValueError('未识别到 Rebrickable parts.csv 所需列（需要 part_num / name / part_cat_id）')

        imported = 0
        with self.lock:
            payload = self._read_rules_payload()
            part_catalog = (
                payload.get('rebrickable_part_catalog')
                if isinstance(payload.get('rebrickable_part_catalog'), dict)
                else {}
            )
            for raw in rows[1:]:
                values = [str(cell or '').strip() for cell in raw]
                if not any(values):
                    continue
                idx = header_map['part_num']
                if idx >= len(values):
                    continue
                part_num = values[idx].strip()
                if not part_num:
                    continue
                name = values[header_map['name']].strip() if header_map['name'] < len(values) else ''
                part_cat_id = values[header_map['part_cat_id']].strip() if header_map['part_cat_id'] < len(values) else ''
                material_idx = header_map.get('part_material')
                part_material = values[material_idx].strip() if material_idx is not None and material_idx < len(values) else ''
                prev = part_catalog.get(part_num) if isinstance(part_catalog.get(part_num), dict) else {}
                part_catalog[part_num] = {
                    'part_num': part_num,
                    'name': name or str(prev.get('name') or '').strip(),
                    'part_cat_id': part_cat_id or str(prev.get('part_cat_id') or '').strip(),
                    'part_material': part_material or str(prev.get('part_material') or '').strip(),
                    'updated_at': _now_iso(),
                    'source_file': safe_name,
                }
                imported += 1
            payload['rebrickable_part_catalog'] = part_catalog
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)
        return {'filename': safe_name, 'imported': imported, 'total_parts': len(part_catalog)}

    def import_rebrickable_part_categories_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        safe_name = str(filename or '').strip()
        rows = self._read_delimited_rows(content)
        if not rows:
            raise ValueError('Rebrickable part_categories.csv 为空或无法解析')
        header = [str(cell or '').strip() for cell in rows[0]]
        header_map = {''.join(str(cell or '').strip().lower().split()): idx for idx, cell in enumerate(header)}
        required = ['id', 'name']
        missing = [key for key in required if key not in header_map]
        if missing:
            raise ValueError('未识别到 Rebrickable part_categories.csv 所需列（需要 id / name）')

        imported = 0
        with self.lock:
            payload = self._read_rules_payload()
            catalog = (
                payload.get('rebrickable_part_category_catalog')
                if isinstance(payload.get('rebrickable_part_category_catalog'), dict)
                else {}
            )
            for raw in rows[1:]:
                values = [str(cell or '').strip() for cell in raw]
                if not any(values):
                    continue
                category_id = values[header_map['id']].strip() if header_map['id'] < len(values) else ''
                if not category_id:
                    continue
                name = values[header_map['name']].strip() if header_map['name'] < len(values) else ''
                catalog[category_id] = {
                    'id': category_id,
                    'name': name,
                    'updated_at': _now_iso(),
                    'source_file': safe_name,
                }
                imported += 1
            payload['rebrickable_part_category_catalog'] = catalog
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)
        return {'filename': safe_name, 'imported': imported, 'total_categories': len(catalog)}

    def import_rebrickable_colors_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        safe_name = str(filename or '').strip()
        rows = self._read_delimited_rows(content)
        if not rows:
            raise ValueError('Rebrickable colors.csv 为空或无法解析')
        header = [str(cell or '').strip() for cell in rows[0]]
        header_map = {''.join(str(cell or '').strip().lower().split()): idx for idx, cell in enumerate(header)}
        required = ['id', 'name', 'rgb', 'is_trans']
        missing = [key for key in required if key not in header_map]
        if missing:
            raise ValueError('未识别到 Rebrickable colors.csv 所需列（需要 id / name / rgb / is_trans）')

        imported = 0
        with self.lock:
            payload = self._read_rules_payload()
            lego_color_catalog = payload.get('lego_color_catalog') if isinstance(payload.get('lego_color_catalog'), dict) else {}
            for raw in rows[1:]:
                values = [str(cell or '').strip() for cell in raw]
                if not any(values):
                    continue
                color_id = values[header_map['id']].strip() if header_map['id'] < len(values) else ''
                if not color_id:
                    continue
                name_en = values[header_map['name']].strip() if header_map['name'] < len(values) else ''
                rgb = values[header_map['rgb']].strip() if header_map['rgb'] < len(values) else ''
                is_trans = values[header_map['is_trans']].strip().lower() == 'true' if header_map['is_trans'] < len(values) else False
                prev = lego_color_catalog.get(color_id) if isinstance(lego_color_catalog.get(color_id), dict) else {}
                name_zh = self._translate_rebrickable_color_name(name_en, str(prev.get('name') or '').strip())
                family = str(prev.get('family') or self._infer_color_family(name_en)).strip() or 'neutral'
                red, green, blue = self._rgb_to_tuple(rgb)
                brightness = self._estimate_brightness(red, green, blue)
                warmth = self._estimate_warmth(red, blue)
                rgb_hex = f'#{rgb.upper()}' if rgb else str(prev.get('rgb_hex') or '').strip()
                lego_color_catalog[color_id] = {
                    'name': name_zh,
                    'name_en': name_en or str(prev.get('name_en') or '').strip(),
                    'family': family,
                    'brightness': brightness if brightness is not None else int(prev.get('brightness', 60) or 60),
                    'warmth': warmth if warmth is not None else int(prev.get('warmth', 50) or 50),
                    'finish': 'transparent' if is_trans else str(prev.get('finish') or 'solid').strip() or 'solid',
                    'rgb_hex': rgb_hex,
                }
                imported += 1
            payload['lego_color_catalog'] = lego_color_catalog
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)
        return {'filename': safe_name, 'imported': imported, 'total_colors': len(lego_color_catalog)}

    def import_rebrickable_elements_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        safe_name = str(filename or '').strip()
        rows = self._read_delimited_rows(content)
        if not rows:
            raise ValueError('Rebrickable elements.csv 为空或无法解析')
        header = [str(cell or '').strip() for cell in rows[0]]
        header_map = {''.join(str(cell or '').strip().lower().split()): idx for idx, cell in enumerate(header)}
        required = ['element_id', 'part_num', 'color_id']
        missing = [key for key in required if key not in header_map]
        if missing:
            raise ValueError('未识别到 Rebrickable elements.csv 所需列（需要 element_id / part_num / color_id）')

        imported = 0
        with self.lock:
            payload = self._read_rules_payload()
            element_map = payload.get('element_part_map') if isinstance(payload.get('element_part_map'), dict) else {}
            for raw in rows[1:]:
                values = [str(cell or '').strip() for cell in raw]
                if not any(values):
                    continue
                element_id = values[header_map['element_id']].strip() if header_map['element_id'] < len(values) else ''
                if not element_id:
                    continue
                part_num = values[header_map['part_num']].strip() if header_map['part_num'] < len(values) else ''
                color_id = values[header_map['color_id']].strip() if header_map['color_id'] < len(values) else ''
                design_idx = header_map.get('design_id')
                design_id = values[design_idx].strip() if design_idx is not None and design_idx < len(values) else ''
                if not part_num:
                    continue
                element_map[element_id] = {
                    'element_id': element_id,
                    'part_num': part_num,
                    'color_id': color_id,
                    'design_id': design_id,
                    'updated_at': _now_iso(),
                    'source_file': safe_name,
                }
                imported += 1
            payload['element_part_map'] = element_map
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)
        return {'filename': safe_name, 'imported': imported, 'total_elements': len(element_map)}

    def import_gobricks_category_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        safe_name = str(filename or '').strip()
        if Path(safe_name).suffix.lower() != '.json':
            raise ValueError('零件分类仅支持 json 文件')
        try:
            data = json.loads(self._decode_text_bytes(content))
        except Exception as exc:
            raise ValueError('零件分类 JSON 无法解析') from exc
        if not isinstance(data, list):
            raise ValueError('零件分类 JSON 必须是数组')

        imported = 0
        with self.lock:
            payload = self._read_rules_payload()
            category_map = (
                payload.get('gobricks_category_index') if isinstance(payload.get('gobricks_category_index'), dict) else {}
            )
            for raw in data:
                item = raw if isinstance(raw, dict) else {}
                category_id = str(item.get('id') or '').strip()
                if not category_id:
                    continue
                category_map[category_id] = {
                    'id': category_id,
                    'index': str(item.get('index') or '').strip(),
                    'title': str(item.get('title') or '').strip(),
                    'title_en': str(item.get('title_en') or '').strip(),
                    'ldd_code': str(item.get('ldd_code') or '').strip(),
                    'is_show': str(item.get('is_show') or '').strip(),
                    'ldd_type': item.get('ldd_type', ''),
                    'src': str(item.get('src') or '').strip(),
                    'updated_at': _now_iso(),
                    'source_file': safe_name,
                }
                imported += 1
            payload['gobricks_category_index'] = category_map
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)
        return {'filename': safe_name, 'imported': imported, 'total_categories': len(category_map)}

    def process_gobricks_conversion_result(
        self,
        source_file: str,
        bom_text: str,
        remote_data: Dict[str, Any],
    ) -> Dict[str, Any]:
        parsed_rows = self._parse_bom_text(bom_text)
        with self.lock:
            rules = self._get_cached_rules_payload()
        exact_part_map = rules.get('exact_part_map') if isinstance(rules.get('exact_part_map'), dict) else {}
        part_alias_map = rules.get('part_alias_map') if isinstance(rules.get('part_alias_map'), dict) else {}
        exact_combo_map = rules.get('exact_combo_map') if isinstance(rules.get('exact_combo_map'), dict) else {}
        gobricks_item_index = (
            rules.get('gobricks_item_index') if isinstance(rules.get('gobricks_item_index'), dict) else {}
        )
        rendered: List[Dict[str, Any]] = []
        rendered_map: Dict[str, Dict[str, Any]] = {}
        seen = set()
        success_entries: List[Dict[str, Any]] = []
        shortage_entries: List[Dict[str, Any]] = []

        def append_row(combo_key: str, payload: Dict[str, Any]) -> None:
            seen.add(combo_key)
            rendered.append(payload)
            rendered_map[combo_key] = payload

        local_cached = 0
        local_printed = 0

        for item in parsed_rows:
            part_no = str(item.get('part_no') or '').strip()
            color_no = str(item.get('color_no') or '').strip()
            combo_key = self._combo_key(part_no, color_no)
            local_exact = self._match_combo_rule(
                exact_combo_map=exact_combo_map,
                part_alias_map=part_alias_map,
                part_no=part_no,
                base_part_no=str(item.get('base_part_no') or part_no or '').strip(),
                color_no=color_no,
            )
            if local_exact:
                local_cached += 1
                append_row(
                    combo_key,
                    {
                        'part_no': part_no,
                        'color_no': color_no,
                        'qty': item.get('qty', 0),
                        'name': item.get('name', ''),
                        'category': 'local_exact_cache',
                        'status_text': '本地缓存命中',
                        'resolved_part_no': str(local_exact.get('gobricks_part_no') or ''),
                        'resolved_color_no': str(local_exact.get('gobricks_color_no') or ''),
                        'detail': '命中历史沉淀的全局映射库',
                    },
                )
                continue
            if str(item.get('tag') or '').strip().lower() == 'printed':
                base_part_no = str(item.get('base_part_no') or part_no or '').strip()
                local_printed += 1
                append_row(
                    combo_key,
                    {
                        'part_no': part_no,
                        'color_no': color_no,
                        'qty': item.get('qty', 0),
                        'name': item.get('name', ''),
                        'category': 'local_printed_fallback',
                        'status_text': '印刷件基础模具回退',
                        'resolved_part_no': f'{base_part_no}-BLANK',
                        'resolved_color_no': color_no,
                        'detail': f'建议先尝试基础模具 {base_part_no} 的空白件/贴纸方案',
                    },
                )

        def row_lookup(part_no: str, color_no: str) -> Dict[str, Any]:
            safe_part = str(part_no or '').strip()
            safe_color = str(color_no or '').strip()
            base_part = self._extract_base_part_no(part_no=safe_part, ldraw_id='')
            exact_candidates = [
                item
                for item in parsed_rows
                if str(item.get('part_no') or '').strip() == safe_part and str(item.get('color_no') or '').strip() == safe_color
            ]
            if exact_candidates:
                return exact_candidates[0]
            exact_part_candidates = [item for item in parsed_rows if str(item.get('part_no') or '').strip() == safe_part]
            if safe_color:
                base_color_candidates = [
                    item
                    for item in parsed_rows
                    if str(item.get('base_part_no') or '').strip() == base_part
                    and str(item.get('color_no') or '').strip() == safe_color
                ]
                if len(base_color_candidates) == 1:
                    return base_color_candidates[0]
            if len(exact_part_candidates) == 1:
                return exact_part_candidates[0]
            base_part_candidates = [item for item in parsed_rows if str(item.get('base_part_no') or '').strip() == base_part]
            if len(base_part_candidates) == 1:
                return base_part_candidates[0]
            if base_part_candidates:
                # Prefer same color within the same base mold; otherwise use the first candidate as fallback context.
                for candidate in base_part_candidates:
                    if str(candidate.get('color_no') or '').strip() == safe_color:
                        return candidate
                return base_part_candidates[0]
            return {
                'part_no': safe_part,
                'color_no': safe_color,
                'qty': 0,
                'name': '',
                'base_part_no': base_part,
                'tag': '',
            }

        def add_shortage(part_no: str, color_no: str, shortage_type: str) -> None:
            shortage_entries.append(
                {
                    'lego_part_no': part_no,
                    'lego_color_no': color_no,
                    'shortage_type': shortage_type,
                    'source_file': str(source_file or '').strip(),
                    'updated_at': _now_iso(),
                }
            )

        def find_index_candidates(part_no: str, color_no: str) -> List[Dict[str, Any]]:
            safe_part = str(part_no or '').strip()
            safe_color = str(color_no or '').strip()
            base_part = self._extract_base_part_no(part_no=safe_part, ldraw_id='')
            return self._find_gobricks_index_candidates(
                gobricks_item_index=gobricks_item_index,
                part_alias_map=part_alias_map,
                part_no=safe_part,
                base_part_no=base_part,
                color_no=safe_color,
            )

        def format_index_suggestions(part_no: str, color_no: str) -> str:
            suggestions = find_index_candidates(part_no, color_no)
            return self._format_gobricks_index_suggestions(suggestions)

        def build_info_detail(info: Dict[str, Any]) -> str:
            if not info:
                return '高砖已返回可售零件'
            chunks = []
            caption = str(info.get('caption') or info.get('caption_en') or '').strip()
            if caption:
                chunks.append(caption)
            shelf_state = str(info.get('shelf_state') or '').strip()
            if shelf_state:
                chunks.append(f'状态 {shelf_state}')
            product_weight = str(info.get('product_weight') or '').strip()
            if product_weight:
                chunks.append(f'重量 {product_weight}g')
            lego_color_id = str(info.get('lego_color_id') or '').strip()
            if lego_color_id:
                chunks.append(f'LEGO色 {lego_color_id}')
            return '；'.join(chunks) if chunks else '高砖已返回可售零件'

        item_list = remote_data.get('itemList') if isinstance(remote_data.get('itemList'), list) else []
        for raw in item_list:
            item = raw if isinstance(raw, dict) else {}
            info = item.get('info') if isinstance(item.get('info'), dict) else {}
            part_no = str(item.get('lego_id') or info.get('lego_id') or item.get('designid') or item.get('design_id') or '').strip()
            color_no = str(
                info.get('lego_color_id')
                or item.get('lego_color_id')
                or item.get('design_color')
                or item.get('colorid')
                or ''
            ).strip()
            combo_key = self._combo_key(part_no, color_no)
            if combo_key in seen:
                continue
            target = self._extract_gobricks_target(
                item_id=str(item.get('item_id') or ''),
                product_id=str(info.get('product_id') or ''),
                color_id=str(info.get('color_id') or ''),
            )
            gds_part_no = str(target.get('part_no') or '').strip()
            gds_color_no = str(target.get('color_no') or '').strip()
            if not gds_part_no:
                continue
            source_row = row_lookup(part_no, color_no)
            success_entries.append(
                {
                    'lego_part_no': part_no,
                    'lego_color_no': color_no,
                    'gobricks_part_no': gds_part_no,
                    'gobricks_color_no': gds_color_no,
                    'source_file': str(source_file or '').strip(),
                    'updated_at': _now_iso(),
                }
            )
            append_row(
                combo_key,
                {
                    'part_no': part_no,
                    'color_no': color_no,
                    'qty': source_row.get('qty', 0),
                    'name': source_row.get('name', '') or str(info.get('caption') or info.get('caption_en') or ''),
                    'category': 'remote_success',
                    'status_text': '高砖接口成功',
                    'resolved_part_no': gds_part_no,
                    'resolved_color_no': gds_color_no,
                    'detail': build_info_detail(info),
                },
            )

        color_deficiency = remote_data.get('colorDeficiency') if isinstance(remote_data.get('colorDeficiency'), list) else []
        for raw in color_deficiency:
            item = raw if isinstance(raw, dict) else {}
            info = item.get('info') if isinstance(item.get('info'), dict) else {}
            part_no = str(item.get('lego_id') or info.get('lego_id') or item.get('designid') or item.get('design_id') or '').strip()
            color_no = str(
                info.get('lego_color_id')
                or item.get('lego_color_id')
                or item.get('design_color')
                or item.get('colorid')
                or ''
            ).strip()
            combo_key = self._combo_key(part_no, color_no)
            if combo_key in seen:
                continue
            source_row = row_lookup(part_no, color_no)
            miss_color = item.get('missColor') if isinstance(item.get('missColor'), dict) else {}
            target = self._extract_gobricks_target(
                item_id=str(item.get('item_id') or ''),
                product_id=str(info.get('product_id') or ''),
                color_id=str(info.get('color_id') or ''),
            )
            detail_parts = ['已识别到同模具，但当前颜色不可售']
            if miss_color:
                detail_parts.append(f"缺色提示：{miss_color.get('name', '')}（高砖色号 {miss_color.get('id', '')}）")
            suggestion = format_index_suggestions(part_no, color_no)
            if suggestion:
                detail_parts.append(suggestion)
            detail = '；'.join(part for part in detail_parts if part)
            add_shortage(part_no, color_no, f'缺少颜色|{detail}')
            append_row(
                combo_key,
                {
                    'part_no': part_no,
                    'color_no': color_no,
                    'qty': source_row.get('qty', 0),
                    'name': source_row.get('name', ''),
                    'category': 'remote_color_deficiency',
                    'status_text': '缺颜色',
                    'resolved_part_no': str(target.get('part_no') or ''),
                    'resolved_color_no': str(target.get('color_no') or ''),
                    'detail': detail,
                },
            )

        inventory_deficiency = (
            remote_data.get('inventoryDeficiency') if isinstance(remote_data.get('inventoryDeficiency'), list) else []
        )
        for raw in inventory_deficiency:
            item = raw if isinstance(raw, dict) else {}
            info = item.get('info') if isinstance(item.get('info'), dict) else {}
            part_no = str(item.get('lego_id') or info.get('lego_id') or item.get('designid') or item.get('design_id') or '').strip()
            color_no = str(
                info.get('lego_color_id')
                or item.get('lego_color_id')
                or item.get('design_color')
                or item.get('colorid')
                or ''
            ).strip()
            combo_key = self._combo_key(part_no, color_no)
            if combo_key in seen:
                continue
            source_row = row_lookup(part_no, color_no)
            target = self._extract_gobricks_target(
                item_id=str(item.get('item_id') or ''),
                product_id=str(info.get('product_id') or ''),
                color_id=str(info.get('color_id') or ''),
            )
            detail_parts = [f"库存不足，可售 {item.get('inventory', 0)}"]
            suggestion = format_index_suggestions(part_no, color_no)
            if suggestion:
                detail_parts.append(suggestion)
            detail = '；'.join(part for part in detail_parts if part)
            add_shortage(part_no, color_no, f'库存不足|{detail}')
            append_row(
                combo_key,
                {
                    'part_no': part_no,
                    'color_no': color_no,
                    'qty': source_row.get('qty', 0),
                    'name': source_row.get('name', ''),
                    'category': 'remote_inventory_deficiency',
                    'status_text': '库存不足',
                    'resolved_part_no': str(target.get('part_no') or ''),
                    'resolved_color_no': str(target.get('color_no') or ''),
                    'detail': detail,
                },
            )

        for group_name, status_text in (
            ('missList', '缺零件'),
            ('noSellList', '不售卖'),
            ('buyLimitList', '限购'),
        ):
            group = remote_data.get(group_name) if isinstance(remote_data.get(group_name), list) else []
            for raw in group:
                item = raw if isinstance(raw, dict) else {}
                info = item.get('info') if isinstance(item.get('info'), dict) else {}
                part_no = str(
                    item.get('designid') or item.get('design_id') or item.get('lego_id') or info.get('lego_id') or ''
                ).strip()
                color_no = str(
                    info.get('lego_color_id')
                    or item.get('lego_color_id')
                    or item.get('design_color')
                    or item.get('colorid')
                    or ''
                ).strip()
                source_row = row_lookup(part_no, color_no)
                row_part_no = str(source_row.get('part_no') or part_no).strip()
                row_color_no = str(source_row.get('color_no') or color_no).strip()
                combo_key = self._combo_key(row_part_no, row_color_no)
                if combo_key in seen:
                    existing = rendered_map.get(combo_key)
                    if isinstance(existing, dict) and str(existing.get('category') or '') == 'local_printed_fallback':
                        effective_existing_status = status_text
                        if status_text == '缺零件':
                            shortage_fix = self._classify_remote_shortage(
                                exact_part_map=exact_part_map,
                                exact_combo_map=exact_combo_map,
                                part_alias_map=part_alias_map,
                                gobricks_item_index=gobricks_item_index,
                                part_no=row_part_no,
                                base_part_no=str(source_row.get('base_part_no') or row_part_no or '').strip(),
                                color_no=row_color_no,
                            )
                            effective_existing_status = str(shortage_fix.get('status_text') or status_text)
                            if str(shortage_fix.get('resolved_part_no') or '').strip():
                                existing['resolved_part_no'] = str(shortage_fix.get('resolved_part_no') or '').strip()
                            if str(shortage_fix.get('resolved_color_no') or '').strip():
                                existing['resolved_color_no'] = str(shortage_fix.get('resolved_color_no') or '').strip()
                        existing['status_text'] = f'印刷件回退 + {effective_existing_status}'
                        extra_detail = []
                        if str(existing.get('detail') or '').strip():
                            extra_detail.append(str(existing.get('detail') or '').strip())
                        extra_detail.append(f'高砖结果：{effective_existing_status}')
                        if str(item.get('info') or '').strip():
                            extra_detail.append(str(item.get('info') or '').strip())
                        existing['detail'] = '；'.join(extra_detail)
                    continue
                effective_status_text = status_text
                resolved_part_no = ''
                resolved_color_no = ''
                detail_parts = []
                if status_text == '缺零件':
                    shortage_fix = self._classify_remote_shortage(
                        exact_part_map=exact_part_map,
                        exact_combo_map=exact_combo_map,
                        part_alias_map=part_alias_map,
                        gobricks_item_index=gobricks_item_index,
                        part_no=row_part_no,
                        base_part_no=str(source_row.get('base_part_no') or row_part_no or '').strip(),
                        color_no=row_color_no,
                    )
                    effective_status_text = str(shortage_fix.get('status_text') or status_text)
                    resolved_part_no = str(shortage_fix.get('resolved_part_no') or '').strip()
                    resolved_color_no = str(shortage_fix.get('resolved_color_no') or '').strip()
                    if str(shortage_fix.get('detail_prefix') or '').strip():
                        detail_parts.append(str(shortage_fix.get('detail_prefix') or '').strip())
                if str(source_row.get('tag') or '').strip().lower() == 'printed':
                    base_part_no = str(source_row.get('base_part_no') or row_part_no or '').strip()
                    detail_parts.append(f'印刷件建议先回退基础模具 {base_part_no} 的空白件/贴纸方案')
                suggestion = format_index_suggestions(row_part_no, row_color_no)
                if suggestion:
                    detail_parts.append(suggestion)
                detail = '；'.join(detail_parts) if detail_parts else str(item.get('info') or effective_status_text)
                add_shortage(row_part_no, row_color_no, f'{effective_status_text}|{detail}')
                append_row(
                    combo_key,
                    {
                        'part_no': row_part_no,
                        'color_no': row_color_no,
                        'qty': source_row.get('qty', 0),
                        'name': source_row.get('name', ''),
                        'category': f'remote_{group_name.lower()}',
                        'status_text': effective_status_text,
                        'resolved_part_no': resolved_part_no,
                        'resolved_color_no': resolved_color_no,
                        'detail': detail,
                    },
                )

        # Persist learning from remote conversion into global rule maps.
        with self.lock:
            payload = self._read_rules_payload()
            exact_combo_map_live = payload.get('exact_combo_map') if isinstance(payload.get('exact_combo_map'), dict) else {}
            shortage_combo_map_live = payload.get('shortage_combo_map') if isinstance(payload.get('shortage_combo_map'), dict) else {}
            for entry in success_entries:
                combo_key = self._combo_key(entry.get('lego_part_no', ''), entry.get('lego_color_no', ''))
                exact_combo_map_live[combo_key] = entry
            for entry in shortage_entries:
                combo_key = self._combo_key(entry.get('lego_part_no', ''), entry.get('lego_color_no', ''))
                shortage_combo_map_live[combo_key] = entry
            payload['exact_combo_map'] = exact_combo_map_live
            payload['shortage_combo_map'] = shortage_combo_map_live
            payload['updated_at'] = _now_iso()
            self._write_json(RULES_PATH, payload)

        corrected_color_deficiency = sum(1 for item in rendered if str(item.get('status_text') or '').strip() == '缺颜色')
        corrected_inventory_deficiency = sum(1 for item in rendered if str(item.get('status_text') or '').strip() == '库存不足')
        corrected_miss = sum(1 for item in rendered if str(item.get('status_text') or '').strip() == '缺零件')
        summary = {
            'input_rows': len(parsed_rows),
            'local_cached_hits': local_cached,
            'local_printed_fallbacks': local_printed,
            'remote_success': len(item_list),
            'remote_color_deficiency': corrected_color_deficiency,
            'remote_inventory_deficiency': corrected_inventory_deficiency,
            'remote_miss': corrected_miss,
            'remote_no_sell': len(remote_data.get('noSellList') or []),
            'remote_buy_limit': len(remote_data.get('buyLimitList') or []),
            'indexed_candidates': sum(1 for item in rendered if '可参考高砖在售同模具候选：' in str(item.get('detail') or '')),
            'rendered_rows': len(rendered),
        }
        return {'summary': summary, 'items': self._decorate_result_rows(results=rendered, rules=rules)}

    def _decorate_result_rows(self, results: List[Dict[str, Any]], rules: Dict[str, Any]) -> List[Dict[str, Any]]:
        source_parts: List[str] = []
        for item in results:
            row = item if isinstance(item, dict) else {}
            part_no = str(row.get('part_no') or '').strip()
            base_part_no = str(row.get('base_part_no') or self._extract_base_part_no(part_no=part_no, ldraw_id='') or '').strip()
            if part_no:
                source_parts.append(part_no)
            if base_part_no:
                source_parts.append(base_part_no)
        detail_map = self._ensure_part_details(source_parts, rules)
        decorated: List[Dict[str, Any]] = []
        for item in results:
            row = deepcopy(item if isinstance(item, dict) else {})
            source_part_no = str(row.get('part_no') or '').strip()
            source_color_no = str(row.get('color_no') or '').strip()
            base_part_no = str(row.get('base_part_no') or self._extract_base_part_no(part_no=source_part_no, ldraw_id='') or '').strip()
            source_name = str(row.get('name') or '').strip()
            local_part_catalog = (
                rules.get('rebrickable_part_catalog') if isinstance(rules.get('rebrickable_part_catalog'), dict) else {}
            )
            local_category_catalog = (
                rules.get('rebrickable_part_category_catalog')
                if isinstance(rules.get('rebrickable_part_category_catalog'), dict)
                else {}
            )
            local_part_meta = (
                local_part_catalog.get(source_part_no) if isinstance(local_part_catalog.get(source_part_no), dict) else {}
            )
            source_detail = detail_map.get(source_part_no) if isinstance(detail_map.get(source_part_no), dict) else {}
            base_detail = detail_map.get(base_part_no) if isinstance(detail_map.get(base_part_no), dict) else {}
            if not source_name or source_name == source_part_no:
                catalog_name = str(local_part_meta.get('name') or '').strip()
                if catalog_name:
                    row['name'] = catalog_name
                    source_name = catalog_name
                fetched_name = str(source_detail.get('name') or '').strip()
                if fetched_name and (not source_name or source_name == source_part_no):
                    row['name'] = fetched_name
                    source_name = fetched_name
            raw_tag = str(row.get('tag') or '').strip().lower()
            tag = raw_tag or self._infer_tag(part_no=source_part_no, part_name=source_name, ldraw_id='')
            is_printed = tag == 'printed'
            exact_image_url = str(source_detail.get('image_url') or '').strip()
            base_image_url = ''
            if not is_printed and base_part_no and base_part_no != source_part_no:
                base_image_url = str(base_detail.get('image_url') or '').strip()
            bricklink_exact_url = self._bricklink_part_image_url(source_part_no, source_color_no, exact=True)
            bricklink_base_url = ''
            if not is_printed and base_part_no and base_part_no != source_part_no:
                bricklink_base_url = self._bricklink_part_image_url(base_part_no, source_color_no, exact=False)
            bricklink_part_url = self._bricklink_part_image_url(source_part_no, '', exact=False)
            ordered_urls = self._ordered_unique_strings(
                [
                    exact_image_url,
                    bricklink_exact_url,
                    bricklink_part_url,
                    base_image_url if not is_printed else '',
                    bricklink_base_url if not is_printed else '',
                ]
            )
            if ordered_urls:
                row['source_image_url'] = ordered_urls[0]
                row['source_image_fallbacks'] = ordered_urls[1:]
                if exact_image_url and ordered_urls[0] == exact_image_url:
                    row['source_image_kind'] = 'rebrickable_exact'
                elif bricklink_exact_url and ordered_urls[0] == bricklink_exact_url:
                    row['source_image_kind'] = 'bricklink_exact'
                elif base_image_url and ordered_urls[0] == base_image_url:
                    row['source_image_kind'] = 'rebrickable_base'
                elif bricklink_base_url and ordered_urls[0] == bricklink_base_url:
                    row['source_image_kind'] = 'bricklink_base'
                else:
                    row['source_image_kind'] = 'bricklink_exact'
            else:
                row['source_image_kind'] = 'placeholder'
                row['source_image_fallbacks'] = []
            row['tag'] = tag
            row['display_group'] = 1 if is_printed else 0
            part_cat_id = str(local_part_meta.get('part_cat_id') or '').strip()
            category_meta = (
                local_category_catalog.get(part_cat_id)
                if part_cat_id and isinstance(local_category_catalog.get(part_cat_id), dict)
                else {}
            )
            row['source_catalog_meta'] = {
                'part_cat_id': part_cat_id,
                'part_cat_name': str(category_meta.get('name') or '').strip(),
                'part_material': str(local_part_meta.get('part_material') or '').strip(),
            }
            row['resolved_meta'] = self._build_resolved_meta(
                resolved_part_no=str(row.get('resolved_part_no') or '').strip(),
                resolved_color_no=str(row.get('resolved_color_no') or '').strip(),
                rules=rules,
            )
            decorated.append(row)
        return decorated

    def _build_resolved_meta(
        self,
        resolved_part_no: str,
        resolved_color_no: str,
        rules: Dict[str, Any],
    ) -> Dict[str, Any]:
        safe_part = str(resolved_part_no or '').strip()
        safe_color = str(resolved_color_no or '').strip()
        if not safe_part or safe_part in {'-', ''}:
            return {}
        if not safe_part.startswith('GDS-'):
            return {}

        item_index = rules.get('gobricks_item_index') if isinstance(rules.get('gobricks_item_index'), dict) else {}
        category_index = (
            rules.get('gobricks_category_index') if isinstance(rules.get('gobricks_category_index'), dict) else {}
        )
        item_id = safe_part if safe_part.count('-') >= 2 else f'{safe_part}-{safe_color}'.rstrip('-')
        entry = item_index.get(item_id) if isinstance(item_index.get(item_id), dict) else {}
        if not entry and safe_part.count('-') >= 2:
            # If part number already contains color, use it directly.
            entry = item_index.get(safe_part) if isinstance(item_index.get(safe_part), dict) else {}

        category_name = str(entry.get('category_name') or '').strip()
        category_icon_url = ''
        if category_name:
            for raw in category_index.values():
                category = raw if isinstance(raw, dict) else {}
                if str(category.get('title') or '').strip() == category_name:
                    category_icon_url = str(category.get('src') or '').strip()
                    if category_icon_url:
                        break

        return {
            'item_id': str(entry.get('item_id') or item_id).strip(),
            'caption': str(entry.get('caption') or '').strip(),
            'inventory': entry.get('inventory', 0),
            'price': entry.get('price', 0),
            'status': str(entry.get('status') or '').strip(),
            'category_name': category_name,
            'category_path': str(entry.get('category_path') or '').strip(),
            'image_url': str(entry.get('image_url') or '').strip(),
            'image_kind': 'part' if str(entry.get('image_url') or '').strip() else 'placeholder',
            'category_icon_url': category_icon_url,
        }

    def update_review_status(self, job_id: str, line_no: int, review_status: str) -> Optional[Dict[str, Any]]:
        safe_job_id = str(job_id or '').strip()
        safe_line_no = int(line_no or 0)
        safe_status = str(review_status or '').strip().lower()
        if not safe_job_id or safe_line_no <= 0:
            return None
        if safe_status not in {'auto_pass', 'pending_review', 'approved', 'rejected', 'hold'}:
            safe_status = 'pending_review'
        with self.lock:
            payload = self._read_json(JOBS_PATH, DEFAULT_JOBS)
            items = payload.get('items')
            source = items if isinstance(items, list) else []
            found: Optional[Dict[str, Any]] = None
            for item in source:
                if str(item.get('job_id') or '') != safe_job_id:
                    continue
                result_list = item.get('results')
                rows = result_list if isinstance(result_list, list) else []
                for row in rows:
                    if int(row.get('line_no') or 0) != safe_line_no:
                        continue
                    row['review_status'] = safe_status
                    break
                item['summary'] = self._build_summary(rows)
                item['updated_at'] = _now_iso()
                found = deepcopy(item)
                break
            if found is None:
                return None
            payload['items'] = source
            self._write_json(JOBS_PATH, payload)
        return found

    def export_job_csv(self, job_id: str) -> Optional[str]:
        job = self.get_job(job_id)
        if not job:
            return None
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            [
                'gaozhuan_part_no',
                'gaozhuan_color_no',
                'qty',
                'source_part_no',
                'source_color_no',
                'match_stage',
                'risk_level',
                'review_status',
                'optimizer_source',
                'optimizer_score',
                'unit_cost',
                'line_cost',
                'note',
            ]
        )
        for row in job.get('results') or []:
            if str(row.get('resolved_part_no') or '').strip() in {'', '-'}:
                continue
            writer.writerow(
                [
                    row.get('resolved_part_no', ''),
                    row.get('resolved_color_no', ''),
                    row.get('qty', 0),
                    row.get('part_no', ''),
                    row.get('color_no', ''),
                    row.get('stage', ''),
                    row.get('risk', ''),
                    row.get('review_status', ''),
                    row.get('candidate_source', ''),
                    row.get('optimizer', {}).get('total_score', '') if isinstance(row.get('optimizer'), dict) else '',
                    row.get('unit_cost', 0),
                    row.get('line_cost', 0),
                    row.get('note', ''),
                ]
            )
        return output.getvalue()

    def import_bom_file(self, filename: str, content: bytes) -> Dict[str, Any]:
        safe_name = str(filename or '').strip()
        suffix = Path(safe_name).suffix.lower()
        if suffix not in {'.csv', '.txt', '.tsv', '.xlsx'}:
            raise ValueError('仅支持 csv / txt / tsv / xlsx 文件')
        if not content:
            raise ValueError('上传文件为空')
        rows = self._parse_tabular_file(filename=safe_name, content=content)
        if not rows:
            raise ValueError('未解析到有效 BOM 行，请检查文件内容')
        bom_lines = []
        preview = []
        for item in rows:
            bom_line = '|'.join(
                [
                    str(item.get('part_no') or '').strip(),
                    str(item.get('color_no') or '').strip(),
                    str(item.get('qty') or 1),
                    str(item.get('name') or '').strip(),
                    str(item.get('tag') or '').strip(),
                    str(item.get('color_name') or '').strip(),
                ]
            )
            bom_lines.append(bom_line.rstrip('|'))
            if len(preview) < 10:
                preview.append(item)
        return {
            'filename': safe_name,
            'row_count': len(rows),
            'bom_text': '\n'.join(bom_lines),
            'preview': preview,
        }

    def _parse_bom_text(self, bom_text: str) -> List[Dict[str, Any]]:
        result: List[Dict[str, Any]] = []
        lines = str(bom_text or '').splitlines()
        for index, raw in enumerate(lines, start=1):
            line = str(raw or '').strip()
            if not line:
                continue
            parts = [segment.strip() for segment in line.split('|')]
            part_no = parts[0] if parts else ''
            color_no = parts[1] if len(parts) > 1 else ''
            qty_value = 1
            try:
                qty_value = int(float(parts[2])) if len(parts) > 2 and parts[2] else 1
            except Exception:
                qty_value = 1
            name = parts[3] if len(parts) > 3 else ''
            raw_tag = parts[4].lower() if len(parts) > 4 else ''
            color_name = parts[5] if len(parts) > 5 else ''
            inferred_tag = raw_tag or self._infer_tag(part_no=part_no, part_name=name, ldraw_id='')
            result.append(
                {
                    'line_no': index,
                    'part_no': part_no,
                    'color_no': color_no,
                    'color_name': color_name,
                    'qty': max(1, qty_value),
                    'name': name,
                    'tag': inferred_tag,
                    'base_part_no': self._extract_base_part_no(part_no=part_no, ldraw_id=''),
                }
            )
        return [item for item in result if str(item.get('part_no') or '').strip()]

    def _resolve_row(
        self,
        item: Dict[str, Any],
        color_mode: str,
        optimizer_mode: str,
        allow_display_sub: bool,
        allow_structural_sub: bool,
        rules: Dict[str, Any],
        analysis_cache: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        exact_part_map = rules.get('exact_part_map') if isinstance(rules.get('exact_part_map'), dict) else {}
        part_alias_map = rules.get('part_alias_map') if isinstance(rules.get('part_alias_map'), dict) else {}
        color_rules = rules.get('color_rules') if isinstance(rules.get('color_rules'), dict) else {}
        substitutions = rules.get('substitutions') if isinstance(rules.get('substitutions'), dict) else {}
        part_meta = rules.get('part_meta') if isinstance(rules.get('part_meta'), dict) else {}
        lego_color_catalog = rules.get('lego_color_catalog') if isinstance(rules.get('lego_color_catalog'), dict) else {}

        part_no = str(item.get('part_no') or '').strip()
        color_no = str(item.get('color_no') or '').strip()
        tag = str(item.get('tag') or '').strip().lower()
        base_part_no = str(item.get('base_part_no') or part_no or '').strip()
        exact_combo_map = rules.get('exact_combo_map') if isinstance(rules.get('exact_combo_map'), dict) else {}
        shortage_combo_map = rules.get('shortage_combo_map') if isinstance(rules.get('shortage_combo_map'), dict) else {}
        gobricks_item_index = rules.get('gobricks_item_index') if isinstance(rules.get('gobricks_item_index'), dict) else {}
        meta = part_meta.get(part_no) if isinstance(part_meta.get(part_no), dict) else {}
        is_printed = tag == 'printed' or bool(meta.get('printed'))
        is_structural = tag == 'structural' or bool(meta.get('structural'))
        shortage_hint = self._match_shortage_rule(
            shortage_combo_map=shortage_combo_map,
            part_alias_map=part_alias_map,
            part_no=part_no,
            base_part_no=base_part_no,
            color_no=color_no,
            analysis_cache=analysis_cache,
        )
        shortage_text = str(shortage_hint.get('shortage_type') or '').strip() if isinstance(shortage_hint, dict) else ''
        shortage_is_color = self._shortage_is_color_issue(shortage_text)
        history_candidates = self._find_exact_combo_candidates(
            exact_combo_map=exact_combo_map,
            part_alias_map=part_alias_map,
            part_no=part_no,
            base_part_no=base_part_no,
            color_no=color_no,
            analysis_cache=analysis_cache,
        )
        index_candidates = self._find_gobricks_index_candidates(
            gobricks_item_index=gobricks_item_index,
            part_alias_map=part_alias_map,
            part_no=part_no,
            base_part_no=base_part_no,
            color_no=color_no,
            analysis_cache=analysis_cache,
        )
        candidate_note_parts = []
        history_suggestion = self._format_exact_combo_suggestions(history_candidates)
        if history_suggestion:
            candidate_note_parts.append(history_suggestion)
        index_suggestion = self._format_gobricks_index_suggestions(index_candidates)
        if index_suggestion:
            candidate_note_parts.append(index_suggestion)
        candidate_suggestion = '；'.join(candidate_note_parts)
        color_recommendations = self._build_color_recommendations(
            lego_color_catalog=lego_color_catalog,
            source_color_no=color_no,
            history_candidates=history_candidates,
            index_candidates=index_candidates,
            analysis_cache=analysis_cache,
        )
        color_recommendation_text = self._format_color_recommendations(color_recommendations)

        candidates: List[Dict[str, Any]] = []

        def add_candidate(
            candidate_source: str,
            stage: str,
            risk: str,
            note: str,
            resolved_part_no: str,
            resolved_color_no: str,
            color_score: int,
            shape_score: int,
        ) -> None:
            candidates.append(
                self._build_optimizer_candidate(
                    item=item,
                    candidate_source=candidate_source,
                    stage=stage,
                    risk=risk,
                    note=note,
                    resolved_part_no=resolved_part_no,
                    resolved_color_no=resolved_color_no,
                    rules=rules,
                    color_score=color_score,
                    shape_score=shape_score,
                )
            )

        combo_exact = self._match_combo_rule(
            exact_combo_map=exact_combo_map,
            part_alias_map=part_alias_map,
            part_no=part_no,
            base_part_no=base_part_no,
            color_no=color_no,
            analysis_cache=analysis_cache,
        )
        if combo_exact:
            add_candidate(
                candidate_source='exact_combo',
                stage='exact_combo_match',
                risk='A',
                note='命中全局高砖映射库',
                resolved_part_no=str(combo_exact.get('gobricks_part_no') or '-'),
                resolved_color_no=str(combo_exact.get('gobricks_color_no') or '-'),
                color_score=100,
                shape_score=98,
            )

        history_exact_part = self._infer_exact_part_from_history(history_candidates)
        exact_part = str(
            exact_part_map.get(self._normalize_part_alias(part_no, part_alias_map))
            or exact_part_map.get(self._normalize_part_alias(base_part_no, part_alias_map))
            or exact_part_map.get(part_no)
            or exact_part_map.get(base_part_no)
            or history_exact_part
            or ''
        ).strip()
        color_group = color_rules.get(color_no) if isinstance(color_rules.get(color_no), dict) else {}
        color_rule = color_group.get(color_mode) if isinstance(color_group, dict) else None
        color_rule_safe = color_rule if isinstance(color_rule, dict) else None

        if is_printed:
            fallback_part_no = str(
                exact_part_map.get(self._normalize_part_alias(base_part_no, part_alias_map))
                or exact_part_map.get(self._normalize_part_alias(part_no, part_alias_map))
                or exact_part_map.get(base_part_no)
                or exact_part_map.get(part_no)
                or ''
            ).strip()
            note_parts = [f'印刷件降级为空白件/贴纸方案，基础模具 {base_part_no or part_no}，需人工确认']
            if shortage_text:
                note_parts.append(f'历史高砖结果：{shortage_text}')
            add_candidate(
                candidate_source='printed_fallback',
                stage='printed_fallback',
                risk='C',
                note='；'.join(note_parts),
                resolved_part_no=fallback_part_no or f'{base_part_no or part_no}-BLANK',
                resolved_color_no=color_no or '-',
                color_score=70,
                shape_score=78,
            )

        if exact_part and color_rule_safe:
            risk = str(color_rule_safe.get('risk') or 'A').upper()
            stage = 'exact_match' if risk == 'A' else 'near_color'
            note = '编号映射 + 颜色直通' if stage == 'exact_match' else str(color_rule_safe.get('note') or '已按近色规则替换')
            add_candidate(
                candidate_source='rule_exact',
                stage=stage,
                risk=risk,
                note=note,
                resolved_part_no=exact_part,
                resolved_color_no=str(color_rule_safe.get('to') or color_no or '-'),
                color_score=100 if risk == 'A' else 82,
                shape_score=95,
            )

        if exact_part and color_rule is None:
            for recommendation in color_recommendations[:2]:
                add_candidate(
                    candidate_source='smart_color',
                    stage='smart_color_match',
                    risk=str(recommendation.get('risk') or 'C').upper(),
                    note=f"近色推荐：LEGO色 {recommendation.get('lego_color_no') or '-'} {recommendation.get('lego_color_name') or ''}，得分 {recommendation.get('score') or 0}",
                    resolved_part_no=str(recommendation.get('resolved_part_no') or exact_part or '-'),
                    resolved_color_no=str(recommendation.get('resolved_color_no') or '-'),
                    color_score=int(recommendation.get('score') or 0),
                    shape_score=95,
                )
            if index_candidates:
                best = index_candidates[0]
                best_target = self._extract_gobricks_target(
                    item_id=str(best.get('item_id') or ''),
                    product_id=str(best.get('product_id') or ''),
                    color_id=str(best.get('color_id') or ''),
                )
                add_candidate(
                    candidate_source='index',
                    stage='indexed_candidate',
                    risk='C',
                    note='本地规则无原色，改用高砖在售同模具候选',
                    resolved_part_no=str(best_target.get('part_no') or '-'),
                    resolved_color_no=str(best_target.get('color_no') or '-'),
                    color_score=int(color_recommendations[0].get('score') or 68) if color_recommendations else 68,
                    shape_score=92,
                )

        sub_group = substitutions.get(part_no) if isinstance(substitutions.get(part_no), dict) else {}
        sub_key = 'structural' if is_structural else 'display'
        selected = sub_group.get(sub_key) if isinstance(sub_group, dict) else None
        sub_rule = selected if isinstance(selected, dict) else None
        sub_enabled = allow_structural_sub if is_structural else allow_display_sub
        if sub_rule and sub_enabled:
            fallback_color = color_rule_safe or {'to': color_no or '-', 'risk': sub_rule.get('risk') or 'B'}
            add_candidate(
                candidate_source='substitution_structural' if is_structural else 'substitution_display',
                stage='substitute',
                risk=str(sub_rule.get('risk') or 'B').upper(),
                note=str(sub_rule.get('note') or '已按相似件规则替换'),
                resolved_part_no=str(sub_rule.get('to') or '-'),
                resolved_color_no=str(fallback_color.get('to') or color_no or '-'),
                color_score=74,
                shape_score=60 if is_structural else 68,
            )
        if sub_rule and not sub_enabled:
            add_candidate(
                candidate_source='blocked',
                stage='blocked_by_policy',
                risk='D',
                note='当前策略未开启该类零件替代',
                resolved_part_no='-',
                resolved_color_no='-',
                color_score=0,
                shape_score=0,
            )

        for entry in history_candidates[:2]:
            add_candidate(
                candidate_source='history_exact',
                stage='indexed_candidate',
                risk='B',
                note=f"历史命中候选：{str(entry.get('gobricks_part_no') or '-')} / {str(entry.get('gobricks_color_no') or '-')}",
                resolved_part_no=str(entry.get('gobricks_part_no') or '-'),
                resolved_color_no=str(entry.get('gobricks_color_no') or '-'),
                color_score=88,
                shape_score=92,
            )
        for index_item in index_candidates[:2]:
            target = self._extract_gobricks_target(
                item_id=str(index_item.get('item_id') or ''),
                product_id=str(index_item.get('product_id') or ''),
                color_id=str(index_item.get('color_id') or ''),
            )
            candidate_note = f"高砖在售候选：{str(index_item.get('item_id') or '-')}"
            if shortage_text:
                candidate_note = f'{candidate_note}；历史高砖结果：{shortage_text}'
            add_candidate(
                candidate_source='index',
                stage='indexed_candidate',
                risk='C',
                note=candidate_note,
                resolved_part_no=str(target.get('part_no') or '-'),
                resolved_color_no=str(target.get('color_no') or '-'),
                color_score=int(color_recommendations[0].get('score') or 68) if color_recommendations else 68,
                shape_score=90,
            )

        if not candidates:
            note_parts = ['未命中编号映射，也没有可用替代件']
            if shortage_is_color:
                note_parts = ['历史高砖结果显示缺颜色，但当前没有可用的同模具候选']
            if shortage_text:
                note_parts.append(f'历史高砖结果：{shortage_text}')
            if candidate_suggestion:
                note_parts.append(candidate_suggestion)
            if color_recommendation_text:
                note_parts.append(color_recommendation_text)
            add_candidate(
                candidate_source='manual',
                stage='manual_color' if shortage_is_color else 'manual_required',
                risk='D',
                note='；'.join(note_parts),
                resolved_part_no='-',
                resolved_color_no='-',
                color_score=0,
                shape_score=0,
            )

        ranked = self._finalize_optimizer_candidates(
            candidates=candidates,
            qty=int(item.get('qty') or 1),
            optimizer_mode=optimizer_mode,
        )
        best = deepcopy(ranked[0]) if ranked else {**item, 'resolved_part_no': '-', 'resolved_color_no': '-', 'stage': 'manual_required', 'risk': 'D', 'note': '未识别到可用方案'}
        best['review_status'] = 'auto_pass' if str(best.get('risk') or '').upper() == 'A' else 'pending_review'
        best['alternatives'] = [
            {
                'resolved_part_no': str(candidate.get('resolved_part_no') or '-'),
                'resolved_color_no': str(candidate.get('resolved_color_no') or '-'),
                'stage': str(candidate.get('stage') or ''),
                'risk': str(candidate.get('risk') or ''),
                'candidate_source': str(candidate.get('candidate_source') or ''),
                'total_score': candidate.get('optimizer', {}).get('total_score', 0) if isinstance(candidate.get('optimizer'), dict) else 0,
                'reason_text': candidate.get('optimizer', {}).get('reason_text', '') if isinstance(candidate.get('optimizer'), dict) else '',
            }
            for candidate in ranked[1:4]
        ]
        if not str(best.get('note') or '').strip():
            best['note'] = str(best.get('optimizer', {}).get('reason_text') or '').strip()
        return best

    def _build_summary(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        total_types = len(results)
        total_qty = sum(max(0, int(item.get('qty') or 0)) for item in results)
        auto_qty = sum(max(0, int(item.get('qty') or 0)) for item in results if str(item.get('risk') or '').upper() == 'A')
        review_qty = sum(max(0, int(item.get('qty') or 0)) for item in results if str(item.get('risk') or '').upper() in {'B', 'C', 'D'})
        blocked_qty = sum(max(0, int(item.get('qty') or 0)) for item in results if str(item.get('risk') or '').upper() == 'D')
        producible_qty = sum(
            max(0, int(item.get('qty') or 0))
            for item in results
            if str(item.get('resolved_part_no') or '').strip() not in {'', '-'}
        )
        estimated_cost_total = round(
            sum(float(item.get('line_cost') or 0) for item in results if float(item.get('line_cost') or 0) > 0),
            2,
        )
        low_stock_count = sum(
            1 for item in results
            if str(item.get('resolved_part_no') or '').strip() not in {'', '-'}
            and int(item.get('inventory') or 0) > 0
            and int(item.get('inventory') or 0) < max(1, int(item.get('qty') or 1))
        )
        unknown_cost_count = sum(
            1 for item in results
            if str(item.get('resolved_part_no') or '').strip() not in {'', '-'}
            and float(item.get('unit_cost') or 0) <= 0
        )
        optimizer_mode = str(results[0].get('optimizer', {}).get('mode') or 'reliability').strip() if results else 'reliability'
        return {
            'total_types': total_types,
            'total_qty': total_qty,
            'auto_qty': auto_qty,
            'review_qty': review_qty,
            'blocked_qty': blocked_qty,
            'producible_qty': producible_qty,
            'auto_match_rate': round((auto_qty / total_qty) * 100, 2) if total_qty else 0,
            'producible_rate': round((producible_qty / total_qty) * 100, 2) if total_qty else 0,
            'optimizer_mode': optimizer_mode,
            'estimated_cost_total': estimated_cost_total,
            'estimated_cost_currency': 'CNY',
            'low_stock_count': low_stock_count,
            'unknown_cost_count': unknown_cost_count,
            'optimized_at': _now_iso(),
        }

    def _combo_key(self, part_no: str, color_no: str) -> str:
        return f'{str(part_no or "").strip()}::{str(color_no or "").strip()}'

    def _normalize_part_alias(self, part_no: str, part_alias_map: Dict[str, Any]) -> str:
        safe_part = str(part_no or '').strip()
        if not safe_part:
            return ''
        mapped = str(part_alias_map.get(safe_part) or '').strip()
        return mapped or safe_part

    def _expand_part_aliases(self, part_no: str, part_alias_map: Dict[str, Any]) -> List[str]:
        safe_part = str(part_no or '').strip()
        if not safe_part:
            return []
        canonical = self._normalize_part_alias(safe_part, part_alias_map)
        candidates: List[str] = [safe_part]
        if canonical and canonical not in candidates:
            candidates.append(canonical)
        for raw_alias, raw_target in part_alias_map.items():
            alias = str(raw_alias or '').strip()
            target = str(raw_target or '').strip()
            if not alias:
                continue
            if alias == canonical or target == canonical:
                if alias not in candidates:
                    candidates.append(alias)
                if target and target not in candidates:
                    candidates.append(target)
        return candidates

    def _expand_part_candidates(
        self,
        part_no: str,
        base_part_no: str,
        part_alias_map: Dict[str, Any],
        analysis_cache: Optional[Dict[str, Any]] = None,
    ) -> List[str]:
        cache = analysis_cache.get('expand_part_candidates_cache') if isinstance(analysis_cache, dict) else None
        cache_key = f'{str(part_no or "").strip()}::{str(base_part_no or "").strip()}'
        if isinstance(cache, dict) and cache_key in cache:
            return list(cache.get(cache_key) or [])
        seen = set()
        ordered: List[str] = []
        for raw in (part_no, base_part_no):
            for candidate in self._expand_part_aliases(raw, part_alias_map):
                if candidate and candidate not in seen:
                    seen.add(candidate)
                    ordered.append(candidate)
        if isinstance(cache, dict):
            cache[cache_key] = list(ordered)
        return ordered

    def _resolve_external_part_no(self, part_no: str, ldraw_id: str, rules: Dict[str, Any]) -> str:
        safe_part = str(part_no or '').strip()
        safe_ldraw = self._normalize_part_no_from_ldraw(ldraw_id)
        external_map = rules.get('external_part_map') if isinstance(rules.get('external_part_map'), dict) else {}
        rebrickable_part_catalog = (
            rules.get('rebrickable_part_catalog') if isinstance(rules.get('rebrickable_part_catalog'), dict) else {}
        )
        for candidate_key in (
            f'part:{safe_part.lower()}',
            f'bricklink:{safe_part.lower()}',
            f'lego:{safe_part.lower()}',
            f'ldraw:{safe_part.lower()}',
            f'ldraw:{safe_ldraw.lower()}',
        ):
            mapped = str(external_map.get(candidate_key) or '').strip()
            if mapped:
                return mapped
        stripped = self._strip_variant_suffix(safe_part)
        if stripped and stripped != safe_part:
            for candidate_key in (
                f'part:{stripped.lower()}',
                f'bricklink:{stripped.lower()}',
                f'lego:{stripped.lower()}',
                f'ldraw:{stripped.lower()}',
            ):
                mapped = str(external_map.get(candidate_key) or '').strip()
                if mapped:
                    return mapped
        if safe_part and safe_part in rebrickable_part_catalog:
            return safe_part
        if stripped and stripped in rebrickable_part_catalog:
            return stripped
        return stripped or safe_part

    def _extract_gobricks_target(
        self,
        item_id: str,
        product_id: str = '',
        color_id: str = '',
    ) -> Dict[str, str]:
        safe_item_id = str(item_id or '').strip()
        safe_product_id = str(product_id or '').strip()
        safe_color_id = str(color_id or '').strip()
        if safe_item_id:
            parts = safe_item_id.split('-')
            if len(parts) >= 3 and parts[0].upper() == 'GDS':
                return {
                    'item_id': safe_item_id,
                    'part_no': '-'.join(parts[:2]),
                    'color_no': parts[2],
                }
        if safe_product_id:
            return {
                'item_id': safe_item_id or f'GDS-{safe_product_id}-{safe_color_id}'.rstrip('-'),
                'part_no': f'GDS-{safe_product_id}',
                'color_no': safe_color_id,
            }
        return {
            'item_id': safe_item_id,
            'part_no': safe_item_id,
            'color_no': safe_color_id,
        }

    def _match_combo_rule(
        self,
        exact_combo_map: Dict[str, Any],
        part_alias_map: Dict[str, Any],
        part_no: str,
        base_part_no: str,
        color_no: str,
        analysis_cache: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        for candidate in self._expand_part_candidates(part_no, base_part_no, part_alias_map, analysis_cache=analysis_cache):
            key = self._combo_key(candidate, color_no)
            value = exact_combo_map.get(key)
            if isinstance(value, dict):
                return value
        return None

    def _infer_exact_part_from_history(
        self,
        history_candidates: List[Dict[str, Any]],
    ) -> str:
        part_nos = {
            str(item.get('gobricks_part_no') or '').strip()
            for item in history_candidates
            if isinstance(item, dict) and str(item.get('gobricks_part_no') or '').strip()
        }
        if len(part_nos) == 1:
            return next(iter(part_nos))
        return ''

    def _match_shortage_rule(
        self,
        shortage_combo_map: Dict[str, Any],
        part_alias_map: Dict[str, Any],
        part_no: str,
        base_part_no: str,
        color_no: str,
        analysis_cache: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        for candidate in self._expand_part_candidates(part_no, base_part_no, part_alias_map, analysis_cache=analysis_cache):
            key = self._combo_key(candidate, color_no)
            value = shortage_combo_map.get(key)
            if isinstance(value, dict):
                return value
        return None

    def _shortage_is_color_issue(self, shortage_text: str) -> bool:
        text = str(shortage_text or '').strip()
        if not text:
            return False
        return '缺少颜色' in text or text.startswith('缺颜色')

    def _find_gobricks_index_candidates(
        self,
        gobricks_item_index: Dict[str, Any],
        part_alias_map: Dict[str, Any],
        part_no: str,
        base_part_no: str,
        color_no: str,
        analysis_cache: Optional[Dict[str, Any]] = None,
    ) -> List[Dict[str, Any]]:
        candidate_parts = set(self._expand_part_candidates(part_no, base_part_no, part_alias_map, analysis_cache=analysis_cache))
        safe_part = str(part_no or '').strip()
        safe_base = str(base_part_no or safe_part or '').strip()
        safe_color = str(color_no or '').strip()
        candidates: List[Dict[str, Any]] = []
        indexed = analysis_cache.get('gobricks_item_part_index') if isinstance(analysis_cache, dict) else None
        if isinstance(indexed, dict):
            seen_ids = set()
            for candidate_part in candidate_parts:
                for raw in indexed.get(candidate_part) or []:
                    item = raw if isinstance(raw, dict) else {}
                    item_key = str(item.get('item_id') or item.get('product_id') or '').strip()
                    if item_key in seen_ids:
                        continue
                    seen_ids.add(item_key)
                    lego_color_id = str(item.get('lego_color_id') or '').strip()
                    if safe_color and lego_color_id and lego_color_id == safe_color:
                        continue
                    candidates.append(item)
        else:
            for raw in gobricks_item_index.values():
                item = raw if isinstance(raw, dict) else {}
                lego_id = str(item.get('lego_id') or '').strip()
                if not lego_id:
                    continue
                lego_base = self._extract_base_part_no(part_no=lego_id, ldraw_id='')
                if lego_id not in candidate_parts and lego_base not in candidate_parts:
                    continue
                lego_color_id = str(item.get('lego_color_id') or '').strip()
                if safe_color and lego_color_id and lego_color_id == safe_color:
                    continue
                candidates.append(item)
        candidates.sort(
            key=lambda item: (
                0 if str(item.get('lego_id') or '').strip() == safe_part else 1,
                0 if str(item.get('lego_id') or '').strip() == safe_base else 1,
                str(item.get('item_id') or ''),
            )
        )
        return candidates[:4]

    def _find_exact_combo_candidates(
        self,
        exact_combo_map: Dict[str, Any],
        part_alias_map: Dict[str, Any],
        part_no: str,
        base_part_no: str,
        color_no: str,
        analysis_cache: Optional[Dict[str, Any]] = None,
    ) -> List[Dict[str, Any]]:
        candidate_parts = set(self._expand_part_candidates(part_no, base_part_no, part_alias_map, analysis_cache=analysis_cache))
        safe_part = str(part_no or '').strip()
        safe_base = str(base_part_no or safe_part or '').strip()
        safe_color = str(color_no or '').strip()
        candidates: List[Dict[str, Any]] = []
        indexed = analysis_cache.get('exact_combo_part_index') if isinstance(analysis_cache, dict) else None
        if isinstance(indexed, dict):
            seen_ids = set()
            for candidate_part in candidate_parts:
                for raw in indexed.get(candidate_part) or []:
                    entry = raw if isinstance(raw, dict) else {}
                    entry_key = self._combo_key(
                        str(entry.get('lego_part_no') or '').strip(),
                        str(entry.get('lego_color_no') or '').strip(),
                    )
                    if entry_key in seen_ids:
                        continue
                    seen_ids.add(entry_key)
                    lego_color_no = str(entry.get('lego_color_no') or '').strip()
                    if safe_color and lego_color_no == safe_color:
                        continue
                    candidates.append(entry)
        else:
            for raw in exact_combo_map.values():
                entry = raw if isinstance(raw, dict) else {}
                lego_part_no = str(entry.get('lego_part_no') or '').strip()
                if not lego_part_no:
                    continue
                lego_base = self._extract_base_part_no(part_no=lego_part_no, ldraw_id='')
                if lego_part_no not in candidate_parts and lego_base not in candidate_parts:
                    continue
                lego_color_no = str(entry.get('lego_color_no') or '').strip()
                if safe_color and lego_color_no == safe_color:
                    continue
                candidates.append(entry)
        candidates.sort(
            key=lambda entry: (
                0 if str(entry.get('lego_part_no') or '').strip() == safe_part else 1,
                0 if str(entry.get('lego_part_no') or '').strip() == safe_base else 1,
                str(entry.get('lego_color_no') or ''),
            )
        )
        return candidates[:4]

    def _format_exact_combo_suggestions(self, candidates: List[Dict[str, Any]]) -> str:
        if not candidates:
            return ''
        chunks = []
        for entry in candidates:
            lego_color_no = str(entry.get('lego_color_no') or '').strip()
            gobricks_part_no = str(entry.get('gobricks_part_no') or '').strip()
            gobricks_color_no = str(entry.get('gobricks_color_no') or '').strip()
            chunks.append(
                f"{gobricks_part_no or '-'} / {gobricks_color_no or '-'}（历史命中 LEGO色 {lego_color_no or '-'}）"
            )
        return '可参考历史成功转换候选：' + '；'.join(chunks)

    def _get_color_meta(self, lego_color_catalog: Dict[str, Any], color_no: str) -> Dict[str, Any]:
        raw = lego_color_catalog.get(str(color_no or '').strip())
        if isinstance(raw, dict):
            return raw
        return {}

    def _score_color_pair(
        self,
        source_meta: Dict[str, Any],
        candidate_meta: Dict[str, Any],
    ) -> int:
        if not source_meta or not candidate_meta:
            return 0
        score = 100
        source_finish = str(source_meta.get('finish') or 'solid').strip()
        candidate_finish = str(candidate_meta.get('finish') or 'solid').strip()
        if source_finish != candidate_finish:
            score -= 35
        source_family = str(source_meta.get('family') or '').strip()
        candidate_family = str(candidate_meta.get('family') or '').strip()
        if source_family != candidate_family:
            score -= 18
        try:
            score -= min(30, abs(int(source_meta.get('brightness', 50)) - int(candidate_meta.get('brightness', 50))))
        except Exception:
            score -= 8
        try:
            score -= min(18, abs(int(source_meta.get('warmth', 50)) - int(candidate_meta.get('warmth', 50))) // 2)
        except Exception:
            score -= 4
        return max(0, min(100, score))

    def _risk_from_color_score(self, score: int) -> str:
        if score >= 85:
            return 'A'
        if score >= 70:
            return 'B'
        return 'C'

    def _build_color_recommendations(
        self,
        lego_color_catalog: Dict[str, Any],
        source_color_no: str,
        history_candidates: List[Dict[str, Any]],
        index_candidates: List[Dict[str, Any]],
        analysis_cache: Optional[Dict[str, Any]] = None,
    ) -> List[Dict[str, Any]]:
        source_meta = self._get_color_meta(lego_color_catalog, source_color_no)
        candidates: List[Dict[str, Any]] = []
        seen_targets = set()

        for entry in history_candidates:
            lego_color_no = str(entry.get('lego_color_no') or '').strip()
            target_part_no = str(entry.get('gobricks_part_no') or '').strip()
            target_color_no = str(entry.get('gobricks_color_no') or '').strip()
            target_key = f'{target_part_no}::{target_color_no}'
            if not target_part_no or target_key in seen_targets:
                continue
            candidate_meta = self._get_color_meta(lego_color_catalog, lego_color_no)
            score = self._score_color_pair_cached(
                source_color_no=source_color_no,
                candidate_color_no=lego_color_no,
                source_meta=source_meta,
                candidate_meta=candidate_meta,
                analysis_cache=analysis_cache,
            )
            candidates.append(
                {
                    'score': score,
                    'risk': self._risk_from_color_score(score),
                    'source': 'history',
                    'lego_color_no': lego_color_no,
                    'lego_color_name': str(candidate_meta.get('name') or '').strip(),
                    'resolved_part_no': target_part_no,
                    'resolved_color_no': target_color_no,
                }
            )
            seen_targets.add(target_key)

        for item in index_candidates:
            target = self._extract_gobricks_target(
                item_id=str(item.get('item_id') or ''),
                product_id=str(item.get('product_id') or ''),
                color_id=str(item.get('color_id') or ''),
            )
            target_part_no = str(target.get('part_no') or '').strip()
            target_color_no = str(target.get('color_no') or '').strip()
            target_key = f'{target_part_no}::{target_color_no}'
            if not target_part_no or target_key in seen_targets:
                continue
            lego_color_no = str(item.get('lego_color_id') or '').strip()
            candidate_meta = self._get_color_meta(lego_color_catalog, lego_color_no)
            score = self._score_color_pair_cached(
                source_color_no=source_color_no,
                candidate_color_no=lego_color_no,
                source_meta=source_meta,
                candidate_meta=candidate_meta,
                analysis_cache=analysis_cache,
            )
            candidates.append(
                {
                    'score': score,
                    'risk': self._risk_from_color_score(score),
                    'source': 'index',
                    'lego_color_no': lego_color_no,
                    'lego_color_name': str(candidate_meta.get('name') or '').strip(),
                    'resolved_part_no': target_part_no,
                    'resolved_color_no': target_color_no,
                }
            )
            seen_targets.add(target_key)

        candidates = [item for item in candidates if int(item.get('score') or 0) >= 60]

        candidates.sort(
            key=lambda item: (
                -int(item.get('score') or 0),
                0 if str(item.get('source') or '') == 'history' else 1,
                str(item.get('resolved_part_no') or ''),
                str(item.get('resolved_color_no') or ''),
            )
        )
        return candidates[:4]

    def _score_color_pair_cached(
        self,
        source_color_no: str,
        candidate_color_no: str,
        source_meta: Dict[str, Any],
        candidate_meta: Dict[str, Any],
        analysis_cache: Optional[Dict[str, Any]] = None,
    ) -> int:
        safe_source = str(source_color_no or '').strip()
        safe_candidate = str(candidate_color_no or '').strip()
        cache = analysis_cache.get('color_pair_score_cache') if isinstance(analysis_cache, dict) else None
        cache_key = f'{safe_source}::{safe_candidate}'
        if isinstance(cache, dict) and cache_key in cache:
            return int(cache.get(cache_key) or 0)
        score = self._score_color_pair(source_meta, candidate_meta)
        if isinstance(cache, dict):
            cache[cache_key] = score
        return score

    def _format_color_recommendations(
        self,
        recommendations: List[Dict[str, Any]],
    ) -> str:
        if not recommendations:
            return ''
        chunks = []
        for item in recommendations:
            chunks.append(
                f"{item.get('resolved_part_no') or '-'} / {item.get('resolved_color_no') or '-'}（评分 {item.get('score') or 0}，风险 {item.get('risk') or '-'}，LEGO色 {item.get('lego_color_no') or '-'} {item.get('lego_color_name') or ''}）"
            )
        return '近色推荐：' + '；'.join(chunks)

    def _format_gobricks_index_suggestions(self, candidates: List[Dict[str, Any]]) -> str:
        if not candidates:
            return ''
        chunks = []
        for item in candidates:
            item_id = str(item.get('item_id') or '').strip()
            caption = str(item.get('caption') or item.get('caption_en') or '').strip()
            lego_color_id = str(item.get('lego_color_id') or '').strip()
            gobricks_color_id = str(item.get('color_id') or '').strip()
            chunks.append(
                f"{item_id or '-'}（LEGO色 {lego_color_id or '-'} -> 高砖色 {gobricks_color_id or '-'}，{caption or '同模具候选'}）"
            )
        return '可参考高砖在售同模具候选：' + '；'.join(chunks)

    def _get_optimizer_weights(self, optimizer_mode: str) -> Dict[str, float]:
        mode = str(optimizer_mode or 'reliability').strip().lower()
        if mode == 'cost':
            return {'success': 0.32, 'availability': 0.18, 'color': 0.12, 'cost': 0.30, 'shape': 0.08}
        if mode == 'appearance':
            return {'success': 0.34, 'availability': 0.16, 'color': 0.28, 'cost': 0.08, 'shape': 0.14}
        return {'success': 0.45, 'availability': 0.20, 'color': 0.15, 'cost': 0.15, 'shape': 0.05}

    def _get_candidate_market_meta(
        self,
        resolved_part_no: str,
        resolved_color_no: str,
        rules: Dict[str, Any],
    ) -> Dict[str, Any]:
        safe_part = str(resolved_part_no or '').strip()
        safe_color = str(resolved_color_no or '').strip()
        if not safe_part or safe_part in {'-', ''}:
            return {}
        item_index = rules.get('gobricks_item_index') if isinstance(rules.get('gobricks_item_index'), dict) else {}
        entry = None
        direct_item_id = f'{safe_part}-{safe_color}'.rstrip('-')
        for key in (direct_item_id, safe_part):
            if isinstance(item_index.get(key), dict):
                entry = item_index.get(key)
                break
        if not isinstance(entry, dict):
            for raw in item_index.values():
                item = raw if isinstance(raw, dict) else {}
                product_id = str(item.get('product_id') or '').strip()
                color_id = str(item.get('color_id') or '').strip()
                normalized_part = f'G{product_id}' if product_id else ''
                if normalized_part and normalized_part == safe_part and (not safe_color or color_id == safe_color):
                    entry = item
                    break
        if not isinstance(entry, dict):
            return {}
        return {
            'item_id': str(entry.get('item_id') or '').strip(),
            'inventory': int(entry.get('inventory') or 0),
            'price': float(entry.get('price') or 0),
            'status': str(entry.get('status') or '').strip(),
        }

    def _build_optimizer_candidate(
        self,
        item: Dict[str, Any],
        candidate_source: str,
        stage: str,
        risk: str,
        note: str,
        resolved_part_no: str,
        resolved_color_no: str,
        rules: Dict[str, Any],
        color_score: int,
        shape_score: int,
    ) -> Dict[str, Any]:
        market_meta = self._get_candidate_market_meta(
            resolved_part_no=resolved_part_no,
            resolved_color_no=resolved_color_no,
            rules=rules,
        )
        return {
            **item,
            'resolved_part_no': str(resolved_part_no or '-').strip() or '-',
            'resolved_color_no': str(resolved_color_no or '-').strip() or '-',
            'stage': stage,
            'risk': str(risk or 'D').upper(),
            'note': str(note or '').strip(),
            'candidate_source': candidate_source,
            'optimizer_market_meta': market_meta,
            'optimizer_color_score': max(0, min(100, int(color_score or 0))),
            'optimizer_shape_score': max(0, min(100, int(shape_score or 0))),
        }

    def _success_score_for_candidate(self, candidate: Dict[str, Any]) -> int:
        source = str(candidate.get('candidate_source') or '').strip().lower()
        risk = str(candidate.get('risk') or 'D').strip().upper()
        base_map = {
            'exact_combo': 100,
            'rule_exact': 95,
            'history_exact': 88,
            'smart_color': 84,
            'index': 76,
            'printed_fallback': 68,
            'substitution_display': 66,
            'substitution_structural': 58,
            'blocked': 18,
            'manual': 8,
        }
        penalty_map = {'A': 0, 'B': 8, 'C': 18, 'D': 40}
        score = base_map.get(source, 50) - penalty_map.get(risk, 25)
        return max(0, min(100, score))

    def _availability_score_for_candidate(self, inventory: int, qty: int, resolved_part_no: str) -> int:
        if str(resolved_part_no or '').strip() in {'', '-'}:
            return 0
        safe_inventory = max(0, int(inventory or 0))
        safe_qty = max(1, int(qty or 1))
        if safe_inventory >= safe_qty * 3:
            return 100
        if safe_inventory >= safe_qty:
            return 80
        if safe_inventory > 0:
            return 55
        return 18

    def _build_optimizer_reason_text(self, candidate: Dict[str, Any], scores: Dict[str, int]) -> str:
        parts: List[str] = []
        source = str(candidate.get('candidate_source') or '').strip().lower()
        market_meta = candidate.get('optimizer_market_meta') if isinstance(candidate.get('optimizer_market_meta'), dict) else {}
        inventory = int(market_meta.get('inventory') or 0)
        price = float(market_meta.get('price') or 0)
        if source == 'exact_combo':
            parts.append('命中历史稳定映射')
        elif source == 'rule_exact':
            parts.append('命中本地规则映射')
        elif source == 'smart_color':
            parts.append('同模具下按近色与在售情况排序')
        elif source == 'index':
            parts.append('未命中本地规则，转用高砖在售候选')
        elif source.startswith('substitution'):
            parts.append('使用相似件替代方案')
        elif source == 'printed_fallback':
            parts.append('印刷件先降级为空白件方案')
        if scores.get('availability', 0) >= 80:
            parts.append(f'库存可覆盖当前数量（{inventory}）')
        elif inventory > 0:
            parts.append(f'库存偏紧（{inventory}）')
        if price > 0:
            parts.append(f'单件约 ¥{price:.2f}')
        if scores.get('color', 0) >= 85:
            parts.append('颜色接近度高')
        elif scores.get('color', 0) >= 70:
            parts.append('颜色可接受')
        elif str(candidate.get('resolved_part_no') or '').strip() not in {'', '-'}:
            parts.append('颜色需人工确认')
        return '；'.join(parts) or str(candidate.get('note') or '').strip() or '系统综合识别率、库存和成本后给出的首选方案'

    def _finalize_optimizer_candidates(
        self,
        candidates: List[Dict[str, Any]],
        qty: int,
        optimizer_mode: str,
    ) -> List[Dict[str, Any]]:
        if not candidates:
            return []
        weights = self._get_optimizer_weights(optimizer_mode)
        priced = [float(item.get('optimizer_market_meta', {}).get('price') or 0) for item in candidates if float(item.get('optimizer_market_meta', {}).get('price') or 0) > 0]
        min_price = min(priced) if priced else 0
        max_price = max(priced) if priced else 0
        finalized: List[Dict[str, Any]] = []
        for raw in candidates:
            candidate = deepcopy(raw)
            market_meta = candidate.get('optimizer_market_meta') if isinstance(candidate.get('optimizer_market_meta'), dict) else {}
            inventory = int(market_meta.get('inventory') or 0)
            price = float(market_meta.get('price') or 0)
            success_score = self._success_score_for_candidate(candidate)
            availability_score = self._availability_score_for_candidate(inventory, qty, str(candidate.get('resolved_part_no') or ''))
            color_score = int(candidate.get('optimizer_color_score') or 0)
            shape_score = int(candidate.get('optimizer_shape_score') or 0)
            if price <= 0:
                cost_score = 50
            elif max_price > min_price:
                cost_score = int(round(100 - ((price - min_price) / (max_price - min_price)) * 100))
            else:
                cost_score = 100
            total_score = round(
                success_score * weights['success']
                + availability_score * weights['availability']
                + color_score * weights['color']
                + cost_score * weights['cost']
                + shape_score * weights['shape'],
                2,
            )
            scores = {
                'success': success_score,
                'availability': availability_score,
                'color': color_score,
                'cost': cost_score,
                'shape': shape_score,
            }
            reason_tags = []
            if success_score >= 90:
                reason_tags.append('HIGH_CONFIDENCE')
            if availability_score < 60:
                reason_tags.append('LOW_STOCK')
            if color_score >= 85:
                reason_tags.append('NEAR_COLOR')
            if cost_score >= 85:
                reason_tags.append('LOW_COST')
            candidate.update(
                {
                    'inventory': inventory,
                    'unit_cost': round(price, 2) if price > 0 else 0,
                    'line_cost': round(price * max(1, int(qty or 1)), 2) if price > 0 else 0,
                    'optimizer': {
                        'mode': optimizer_mode,
                        'total_score': total_score,
                        'score_breakdown': scores,
                        'candidate_source': str(candidate.get('candidate_source') or '').strip(),
                        'reason_tags': reason_tags,
                        'reason_text': self._build_optimizer_reason_text(candidate, scores),
                    },
                }
            )
            finalized.append(candidate)
        finalized.sort(
            key=lambda item: (
                -float(item.get('optimizer', {}).get('total_score') or 0),
                str(item.get('risk') or 'D'),
                -int(item.get('inventory') or 0),
                float(item.get('unit_cost') or 0) if float(item.get('unit_cost') or 0) > 0 else 999999,
                str(item.get('resolved_part_no') or ''),
                str(item.get('resolved_color_no') or ''),
            )
        )
        for index, item in enumerate(finalized, start=1):
            optimizer = item.get('optimizer') if isinstance(item.get('optimizer'), dict) else {}
            optimizer['rank'] = index
            item['optimizer'] = optimizer
        return finalized

    def _classify_remote_shortage(
        self,
        exact_part_map: Dict[str, Any],
        exact_combo_map: Dict[str, Any],
        part_alias_map: Dict[str, Any],
        gobricks_item_index: Dict[str, Any],
        part_no: str,
        base_part_no: str,
        color_no: str,
    ) -> Dict[str, Any]:
        history_candidates = self._find_exact_combo_candidates(
            exact_combo_map=exact_combo_map,
            part_alias_map=part_alias_map,
            part_no=part_no,
            base_part_no=base_part_no,
            color_no=color_no,
        )
        inferred_part = self._infer_exact_part_from_history(history_candidates)
        mapped_part = str(
            exact_part_map.get(self._normalize_part_alias(part_no, part_alias_map))
            or exact_part_map.get(self._normalize_part_alias(base_part_no, part_alias_map))
            or exact_part_map.get(part_no)
            or exact_part_map.get(base_part_no)
            or inferred_part
            or ''
        ).strip()
        index_candidates = self._find_gobricks_index_candidates(
            gobricks_item_index=gobricks_item_index,
            part_alias_map=part_alias_map,
            part_no=part_no,
            base_part_no=base_part_no,
            color_no=color_no,
        )
        if mapped_part or index_candidates:
            target_part_no = mapped_part
            target_color_no = ''
            if not target_part_no and index_candidates:
                first_target = self._extract_gobricks_target(
                    item_id=str(index_candidates[0].get('item_id') or ''),
                    product_id=str(index_candidates[0].get('product_id') or ''),
                    color_id=str(index_candidates[0].get('color_id') or ''),
                )
                target_part_no = str(first_target.get('part_no') or '').strip()
                target_color_no = str(first_target.get('color_no') or '').strip()
            return {
                'status_text': '缺颜色',
                'detail_prefix': '高砖接口返回缺零件，但本地已识别到同模具，按缺颜色处理',
                'resolved_part_no': target_part_no,
                'resolved_color_no': target_color_no,
                'history_candidates': history_candidates,
                'index_candidates': index_candidates,
            }
        return {
            'status_text': '缺零件',
            'detail_prefix': '',
            'resolved_part_no': '',
            'resolved_color_no': '',
            'history_candidates': history_candidates,
            'index_candidates': index_candidates,
        }

    def _parse_tabular_file(self, filename: str, content: bytes) -> List[Dict[str, Any]]:
        suffix = Path(str(filename or '')).suffix.lower()
        if suffix == '.xlsx':
            rows = self._read_xlsx_rows(content)
        else:
            rows = self._read_delimited_rows(content)
        with self.lock:
            rules = self._get_cached_rules_payload()
        return self._normalize_table_rows(rows, rules=rules)

    def _read_delimited_rows(self, content: bytes) -> List[List[str]]:
        text = self._decode_text_bytes(content)
        if not text.strip():
            return []
        sample = text[:2048]
        delimiter = ','
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            delimiter = dialect.delimiter
        except Exception:
            lines = [line for line in text.splitlines() if line.strip()]
            if lines:
                first = lines[0]
                delimiter_scores = {
                    '|': first.count('|'),
                    '\t': first.count('\t'),
                    ',': first.count(','),
                    ';': first.count(';'),
                }
                delimiter = max(delimiter_scores, key=delimiter_scores.get)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        return [[str(cell or '').strip() for cell in row] for row in reader]

    def _read_xlsx_rows(self, content: bytes) -> List[List[str]]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:
            raise ValueError('缺少 openpyxl 依赖，无法解析 xlsx，请先安装 requirements.txt') from exc
        buffer = io.BytesIO(content)
        workbook = load_workbook(buffer, read_only=True, data_only=True)
        sheet = workbook.active
        rows: List[List[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() if cell is not None else '' for cell in row]
            rows.append(values)
        return rows

    def _decode_text_bytes(self, content: bytes) -> str:
        for encoding in ('utf-8-sig', 'utf-8', 'gb18030', 'gbk'):
            try:
                return content.decode(encoding)
            except Exception:
                continue
        return content.decode('utf-8', errors='ignore')

    def _normalize_table_rows(self, rows: List[List[str]], rules: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        cleaned = []
        for row in rows:
            if not row:
                continue
            values = [str(cell or '').strip() for cell in row]
            if not any(values):
                continue
            cleaned.append(values)
        if not cleaned:
            return []

        header = cleaned[0]
        use_header = self._looks_like_header(header)
        start_index = 1 if use_header else 0
        header_index = self._build_header_index(header) if use_header else {}

        result: List[Dict[str, Any]] = []
        live_rules = rules if isinstance(rules, dict) else self.get_rules()
        for row in cleaned[start_index:]:
            item = self._row_to_bom_item(row=row, header_index=header_index, use_header=use_header, rules=live_rules)
            if item:
                result.append(item)
        return result

    def _pick_from_row(self, row: List[str], header_index: Dict[str, int], field: str) -> str:
        idx = header_index.get(field, -1)
        if idx < 0 or idx >= len(row):
            return ''
        return str(row[idx] or '').strip().lstrip("'")

    def _pick_gobricks_part(self, row: List[str], header_row: List[str]) -> str:
        for idx, cell in enumerate(header_row):
            text = str(cell or '').strip().lower()
            if 'gobrick part' in text:
                return str(row[idx] or '').strip().lstrip("'")
        return ''

    def _pick_gobricks_color(self, row: List[str], header_row: List[str]) -> str:
        for idx, cell in enumerate(header_row):
            text = str(cell or '').strip().lower()
            if 'gobrick color' in text:
                return str(row[idx] or '').strip().lstrip("'")
        return ''

    def _pick_shortage_type(self, row: List[str], header_row: List[str]) -> str:
        for idx, cell in enumerate(header_row):
            text = str(cell or '').strip().lower()
            if text == '类型':
                return str(row[idx] or '').strip()
        return ''

    def _is_valid_gobricks_mapping(self, gobrick_part: str, gobrick_color: str) -> bool:
        part = str(gobrick_part or '').strip().lower()
        color = str(gobrick_color or '').strip().lower()
        invalid = {'', '无', 'none', 'null', 'n/a'}
        if part in invalid or part.endswith('-无'):
            return False
        if color in invalid:
            return False
        return True

    def _looks_like_header(self, row: List[str]) -> bool:
        joined = ' '.join(str(cell or '').strip().lower() for cell in row)
        keywords = ['part', 'part no', 'part_no', 'color', 'qty', 'quantity', '编号', '颜色', '数量', '零件']
        return any(keyword in joined for keyword in keywords)

    def _build_header_index(self, row: List[str]) -> Dict[str, int]:
        index_map: Dict[str, int] = {}
        for idx, cell in enumerate(row):
            value = str(cell or '').strip().lower()
            if not value:
                continue
            normalized = ''.join(ch for ch in value if ch.isalnum())
            if normalized in {
                'part',
                'partno',
                'partnumber',
                'blitemno',
                'blitemid',
                'itemno',
                'itemnumber',
                'bricklinkitemno',
                'legoitemno',
                '零件编号',
                '零件号',
                '编号',
            }:
                index_map['part_no'] = idx
            elif normalized in {
                'color',
                'colorno',
                'colour',
                'colourno',
                'blcolorid',
                'bricklinkcolorid',
                'legocolorid',
                '颜色',
                '颜色编号',
                '色号',
            }:
                index_map['color_no'] = idx
            elif normalized in {'qty', 'quantity', 'count', '数量', '件数'}:
                index_map['qty'] = idx
            elif normalized in {'name', 'partname', '零件名', '零件名称', '名称'}:
                index_map['name'] = idx
            elif normalized in {'colorname', 'colourname', '颜色名', '颜色名称'}:
                index_map['color_name'] = idx
            elif normalized in {'tag', '标签', 'type', '类型'}:
                index_map['tag'] = idx
            elif normalized in {'ldrawid', 'ldrawpartid', 'ldrawno'}:
                index_map['ldraw_id'] = idx
            elif normalized in {'elementid', 'legoelementid', 'elementno'}:
                index_map['element_id'] = idx
        return index_map

    def _row_to_bom_item(
        self,
        row: List[str],
        header_index: Dict[str, int],
        use_header: bool,
        rules: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        live_rules = rules if isinstance(rules, dict) else self.get_rules()
        element_part_map = live_rules.get('element_part_map') if isinstance(live_rules.get('element_part_map'), dict) else {}
        rebrickable_part_catalog = (
            live_rules.get('rebrickable_part_catalog') if isinstance(live_rules.get('rebrickable_part_catalog'), dict) else {}
        )
        print_part_map = live_rules.get('print_part_map') if isinstance(live_rules.get('print_part_map'), dict) else {}

        def pick(field: str, fallback_index: int, allow_fallback_when_header_missing: bool = True) -> str:
            if use_header and field in header_index:
                idx = header_index[field]
            elif use_header and not allow_fallback_when_header_missing:
                idx = -1
            else:
                idx = fallback_index
            if idx < 0 or idx >= len(row):
                return ''
            return str(row[idx] or '').strip()

        part_no = pick('part_no', 0)
        color_no = pick('color_no', 1)
        qty_raw = pick('qty', 2)
        name = pick('name', 3)
        color_name = pick('color_name', 5, True)
        tag = pick('tag', 4, False).lower()
        ldraw_id = pick('ldraw_id', -1, False)
        element_id = pick('element_id', -1, False)

        if not part_no and element_id:
            element_meta = element_part_map.get(element_id) if isinstance(element_part_map.get(element_id), dict) else {}
            mapped_part = str(element_meta.get('part_num') or '').strip()
            mapped_color = str(element_meta.get('color_id') or '').strip()
            if mapped_part:
                part_no = mapped_part
            if not color_no and mapped_color:
                color_no = mapped_color

        if not part_no and ldraw_id:
            part_no = self._normalize_part_no_from_ldraw(ldraw_id)

        if part_no:
            part_no = self._resolve_external_part_no(part_no, ldraw_id, live_rules)

        if part_no and not name:
            part_meta = (
                rebrickable_part_catalog.get(part_no) if isinstance(rebrickable_part_catalog.get(part_no), dict) else {}
            )
            fetched_name = str(part_meta.get('name') or '').strip()
            if fetched_name:
                name = fetched_name

        base_part_no = self._extract_base_part_no(part_no=part_no, ldraw_id=ldraw_id)
        print_base = str(print_part_map.get(part_no) or '').strip()
        if print_base:
            base_part_no = print_base

        if not tag:
            tag = self._infer_tag(part_no=part_no, part_name=name, ldraw_id=ldraw_id)

        if not part_no:
            return None
        try:
            qty = int(float(qty_raw)) if qty_raw else 1
        except Exception:
            qty = 1
        return {
            'part_no': part_no,
            'color_no': color_no,
            'qty': max(1, qty),
            'name': name,
            'color_name': color_name,
            'tag': tag,
            'ldraw_id': ldraw_id,
            'element_id': element_id,
            'base_part_no': base_part_no,
        }

    def _normalize_part_no_from_ldraw(self, raw: str) -> str:
        value = str(raw or '').strip()
        if not value:
            return ''
        lower = value.lower()
        if lower.endswith('.dat'):
            return value[:-4]
        return value

    def _strip_variant_suffix(self, part_no: str) -> str:
        safe = str(part_no or '').strip()
        if not safe:
            return ''
        lower = safe.lower()
        match = re.match(r'^([0-9]+[a-z]?)-f[0-9]+$', lower)
        if match:
            return match.group(1)
        return safe

    def _infer_tag(self, part_no: str, part_name: str, ldraw_id: str) -> str:
        safe_part_no = str(part_no or '').strip().lower()
        safe_name = str(part_name or '').strip().lower()
        safe_ldraw = str(ldraw_id or '').strip().lower()

        if 'sticker' in safe_name:
            return 'printed'
        if 'pattern' in safe_name:
            return 'printed'
        if 'print' in safe_name:
            return 'printed'
        if 'pb' in safe_part_no:
            return 'printed'
        if safe_ldraw.endswith('.dat'):
            stem = safe_ldraw[:-4]
        else:
            stem = safe_ldraw
        if 'p' in stem and any(ch.isdigit() for ch in stem):
            # Studio/LDraw printed variants often use p/d suffixed dat ids.
            if any(token in stem for token in ('pb', 'p0', 'p1', 'p2', 'd0', 'dy', 'pt')):
                return 'printed'
        return ''

    def _extract_base_part_no(self, part_no: str, ldraw_id: str) -> str:
        safe_part = str(part_no or '').strip()
        safe_ldraw = self._normalize_part_no_from_ldraw(ldraw_id)

        for candidate in (safe_ldraw, safe_part):
            value = str(candidate or '').strip().lower()
            if not value:
                continue
            variant = re.match(r'^(\d+[a-z]?)-f[0-9]+$', value)
            if variant:
                return variant.group(1)
            direct = re.match(r'^(\d+[a-z]?)$', value)
            if direct:
                return direct.group(1)
            prefixed = re.match(r'^(\d+[a-z]?)(?:pb[0-9a-z]+|pr[0-9a-z]+|p[0-9a-z]+|d[0-9a-z]+|dy[0-9a-z]+|pt[0-9a-z]+)$', value)
            if prefixed:
                return prefixed.group(1)
            fallback = re.match(r'^(\d+)', value)
            if fallback and fallback.group(1) != value:
                return fallback.group(1)
        return safe_part


part_adapter_store = PartAdapterStore()
